# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import base64
import json
import uuid
import zipfile
from datetime import datetime, date, time, timedelta
from typing import Tuple, Union, Optional
from calendar import monthrange
import calendar
from datetime import datetime, date as date_module
import re
import pytz
import gc
from odoo.exceptions import UserError

from collections import defaultdict
import logging
from PyPDF2 import PdfFileReader, PdfFileWriter

from odoo.http import request, route, Controller, content_disposition
from odoo.tools.safe_eval import safe_eval
from odoo import http
import io
import xlsxwriter
from dateutil.relativedelta import relativedelta


### pdf import
from typing import Dict, List, Generator, Set

import tempfile
import os
from subprocess import run, PIPE, TimeoutExpired

import os
import tempfile
from subprocess import run
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

_logger = logging.getLogger(__name__)

meses_espanol = {
    "January": "enero",
    "February": "febrero",
    "March": "marzo",
    "April": "abril",
    "May": "mayo",
    "June": "junio",
    "July": "julio",
    "August": "agosto",
    "September": "septiembre",
    "October": "octubre",
    "November": "noviembre",
    "December": "diciembre",
}

MESES_ESPANOL = {
    'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
    'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
    'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
}

TRANSLATE_DAYS = {
    'Monday': 'lun', 'Tuesday': 'mar', 'Wednesday': 'mie', 'Thursday': 'jue',
    'Friday': 'vie', 'Saturday': 'sab', 'Sunday': 'dom'
}

HEADERS = [
    'Fecha', 'Día', 'Ent 1', 'Sal 1', 'Ent 2', 'Sal 2', 'Ent 3', 'Sal 3',
    'H Trab.', 'Atras.', 'H 25%', 'H 50%', 'H 100%', 'H Debe', 'Ref. Ubic.', 'Observaciones'
]

class HrPayroll(Controller):

    def __init__(self):
        # Cache system parameter once
        self.type_of_resource = request.env['ir.config_parameter'].sudo().get_param(
            'hr_payroll.mode_of_attendance'
        )




    ######## PDF PARA REPORTE ASISTENCIAS #########

    def process_mega_dataset(self, employee_ids: List[int], date_range: List[datetime]) -> Dict:
        """
        Procesa un dataset masivo de asistencias de forma optimizada.
        """
        pin_to_id = self._build_pin_mapping(employee_ids)
        attendance_stream = self._lazy_load_attendances(pin_to_id.keys(), date_range)

        return self._stream_process_attendances(attendance_stream, pin_to_id)

    def _build_pin_mapping(self, employee_ids: List[int]) -> Dict[str, int]:
        """Construye mapeo PIN -> employee_id de forma eficiente."""
        employees = request.env['hr.employee'].browse(employee_ids)
        return {emp.pin: emp.id for emp in employees if emp.pin}

    def _lazy_load_attendances(self, employee_pins: Set[str], date_range: List[datetime]) -> Generator:

        start_range = self.convert_to_utc(datetime.combine(min(date_range), time.min))
        end_range = self.convert_to_utc(datetime.combine(max(date_range), time.max))

        pin_list = list(employee_pins)

        # Batch size dinámico - ajustar según disponibilidad de memoria
        base_limit = 5000
        limit = min(base_limit, len(pin_list) * 100)
        offset = 0

        domain = [
            ('user_id', 'in', pin_list),
            ('timestamp', '>=', start_range),
            ('timestamp', '<=', end_range),
        ]

        while True:
            batch = request.env['hr.attendance.general'].sudo().search(
                domain,
                limit=limit,
                offset=offset,
                order='user_id, timestamp DESC'
            )

            if not batch:
                break

            # Yield cada record del batch
            yield from batch

            offset += limit

            if len(batch) < limit:
                break

    def _stream_process_attendances(self, attendance_stream: Generator, pin_to_id: Dict[str, int]) -> Dict:
        result = defaultdict(lambda: defaultdict(list))

        for attendance in attendance_stream:
            employee_id = pin_to_id.get(attendance.user_id)
            if employee_id:
                # date_key = attendance.timestamp.date()
                date_key = self.convertir_a_hora_ecuador(attendance.timestamp).date()
                result[employee_id][date_key].append(attendance)

        # Convertir a dict normal y ordenar solo una vez por fecha
        final_result = {}
        for employee_id, dates_data in result.items():
            employee_data = {}
            for date_key, attendances in dates_data.items():
                # Ordenar solo si hay más de un registro (optimización)
                if len(attendances) > 1:
                    employee_data[date_key] = sorted(attendances, key=lambda att: att.timestamp)
                else:
                    employee_data[date_key] = attendances
            final_result[employee_id] = employee_data

        return final_result

    @http.route('/reporte_asistencias/download_pdf/<int:record_id>', type='http', auth='public')
    def download_pdf_of_report(self, record_id, **kwargs):
        # Early validation of record
        record = request.env['report.attendance.general'].sudo().browse(record_id)
        model_import = request.env['hr.attendance.import'].sudo()
        if not record.exists():
            return request.not_found()

        # Use employee_ids directly from record
        employees = record.employee_ids
        if not employees:
            return request.make_response(
                b"No employees found for this report.",
                headers=[('Content-Type', 'text/plain')]
            )

        # Get month name in Spanish
        name_mount = record.date_from.strftime("%B")
        name_mount = MESES_ESPANOL.get(name_mount, name_mount)

        # Prefetch holidays and calendar data
        date_range = [
            record.date_from + timedelta(days=x)
            for x in range((record.date_to - record.date_from).days + 1)
        ]
        date_utc_ranges = [
            (
                record.convert_to_utc(datetime.combine(date, time.min) + timedelta(minutes=1)),
                record.convert_to_utc(datetime.combine(date, time.max) - timedelta(minutes=1))
            )
            for date in date_range
        ]

        holidays_dict = record._prefetch_holidays(employees, date_utc_ranges)
        calendar_dict = model_import.get_range_resource_calendar_massive(employees.ids, date_range,
                                                                         self.type_of_resource)
        holidays_employee_dict = record._prefetch_holidays_employee(employees.ids, date_utc_ranges)
        holidays_employee_permits_dict_all = record._prefetch_holidays_employee_permits_all(employees.ids,
                                                                                            date_utc_ranges)
        schedules_names_by_employee = record.prefetch_calendar_names(employees.ids, record.date_from, record.date_to)

        # Prefetch employee IDs
        employee_ids = employees.ids

        # Define date range in UTC
        date_from = record.convert_to_utc(datetime.combine(record.date_from, time.min))
        date_to = record.convert_to_utc(datetime.combine(record.date_to, time.max))

        work_entries = request.env['hr.work.entry'].sudo().search([
            ('employee_id', 'in', employee_ids),
            ('date_start', '>=', date_from),
            ('date_stop', '<=', date_to),
        ])

        work_attendances = self.process_mega_dataset(employee_ids, date_range)

        try:
            pdf_buffer = io.BytesIO()

            if len(employees) <= 20:
                xlsx_buffer = io.BytesIO()
                try:
                    workbook = xlsxwriter.Workbook(xlsx_buffer, {'in_memory': True})
                    self._generate_workbook_for_group(
                        workbook,
                        employees,
                        record,
                        name_mount,
                        holidays_dict,
                        calendar_dict,
                        holidays_employee_dict,
                        holidays_employee_permits_dict_all,
                        schedules_names_by_employee,
                        date_range=date_range,
                        work_entries=work_entries,
                        model_import=model_import,
                        work_attendances=work_attendances
                    )
                    workbook.close()
                    xlsx_buffer.seek(0)
                    self._convert_excel_to_pdf(xlsx_buffer, pdf_buffer)
                finally:
                    xlsx_buffer.close()
            else:
                employees_per_file = 500
                excel_buffers = []
                total_groups = (len(employees) + employees_per_file - 1) // employees_per_file

                for group_index in range(total_groups):
                    start_idx = group_index * employees_per_file
                    end_idx = min(start_idx + employees_per_file, len(employees))
                    employee_group = employees[start_idx:end_idx]

                    if not employee_group:
                        continue

                    xlsx_buffer = io.BytesIO()
                    try:
                        workbook = xlsxwriter.Workbook(xlsx_buffer, {'in_memory': True})
                        self._generate_workbook_for_group(
                            workbook,
                            employee_group,
                            record,
                            name_mount,
                            holidays_dict,
                            calendar_dict,
                            holidays_employee_dict,
                            holidays_employee_permits_dict_all,
                            schedules_names_by_employee,
                            date_range=date_range,
                            work_entries=work_entries,
                            model_import=model_import,
                            work_attendances=work_attendances
                        )
                        workbook.close()
                        xlsx_buffer.seek(0)
                        excel_buffers.append(xlsx_buffer)
                    except Exception:
                        xlsx_buffer.close()
                        raise

                try:
                    self._convert_multiple_excel_to_pdf(excel_buffers, pdf_buffer)
                finally:
                    for buffer in excel_buffers:
                        buffer.close()

            pdf_buffer.seek(0)
            response = request.make_response(
                pdf_buffer.getvalue(),
                headers=[
                    ('Content-Type', 'application/pdf'),
                    ('Content-Disposition', content_disposition(f'reporte_asistencias_{name_mount}.pdf'))
                ]
            )
            return response

        except Exception as e:
            _logger.error(f"Error generating PDF report: {e}")
            return request.make_response(
                b"Error generating PDF report",
                status=500,
                headers=[('Content-Type', 'text/plain')]
            )
        finally:
            pdf_buffer.close()

    def _convert_excel_to_pdf(self, excel_buffer, pdf_buffer):
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
                temp_excel.write(excel_buffer.getvalue())
                temp_excel_path = temp_excel.name

            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
                temp_pdf_path = temp_pdf.name

            try:
                wb = openpyxl.load_workbook(temp_excel_path)

                # DEFINE LAS COLUMNAS PARA AJUSTAR ALTURA
                columns_adjust_height = ['P']  # Columnas que ajustan altura

                # DEFINE LAS COLUMNAS PARA AJUSTAR ANCHO
                columns_adjust_width = ['P']  # Columnas que autoajustan ancho

                for sheet in wb.worksheets:

                    # PASO 1: Autoajustar ANCHO de columnas específicas
                    for col_letter in columns_adjust_width:
                        max_length = 0

                        for cell in sheet[col_letter]:
                            # SKIP celdas combinadas
                            if isinstance(cell, MergedCell):
                                continue

                            if cell.value:
                                # Habilitar wrap_text
                                if cell.alignment:
                                    cell.alignment = cell.alignment.copy(
                                        wrap_text=True,
                                        vertical='top'
                                    )
                                else:
                                    cell.alignment = Alignment(
                                        wrap_text=True,
                                        vertical='top'
                                    )

                                # Calcular longitud máxima
                                lines = str(cell.value).split('\n')
                                for line in lines:
                                    max_length = max(max_length, len(line))

                        # Aplicar ancho óptimo
                        if max_length > 0:
                            optimal_width = min(max_length * 1.2 + 2, 100)
                            sheet.column_dimensions[col_letter].width = optimal_width

                    # PASO 2: Ajustar ALTURA de filas (considerando columnas específicas)
                    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                        max_height = 28  # Altura mínima

                        for cell in row:
                            # SKIP celdas combinadas
                            if isinstance(cell, MergedCell):
                                continue

                            # SOLO procesar columnas específicas para altura
                            # if cell.column_letter not in columns_adjust_height:
                            #     continue

                            if cell.value:
                                # Asegurar que tenga wrap_text
                                if cell.alignment:
                                    cell.alignment = cell.alignment.copy(
                                        wrap_text=True,
                                        vertical='top'
                                    )
                                else:
                                    cell.alignment = Alignment(
                                        wrap_text=True,
                                        vertical='top'
                                    )

                                # Obtener ancho actual de la columna
                                col_width = sheet.column_dimensions[cell.column_letter].width or 10
                                chars_per_line = max(int((col_width - 2) / 1.2), 1)

                                # Calcular líneas necesarias
                                lines = str(cell.value).split('\n')
                                total_lines = 0

                                for line in lines:
                                    if len(line) == 0:
                                        total_lines += 1
                                    else:
                                        total_lines += (len(line) // chars_per_line) + 1

                                needed_height = total_lines * 5  # Cambiado a 15 (más estándar)
                                max_height = max(max_height, needed_height)

                        # Aplicar altura calculada
                        sheet.row_dimensions[row_idx].height = min(max_height, 300)

                wb.save(temp_excel_path)
                wb.close()

                # Convertir con LibreOffice
                result = run([
                    'libreoffice',
                    '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', os.path.dirname(temp_pdf_path),
                    temp_excel_path
                ], capture_output=True, text=True, timeout=30)

                if result.returncode != 0:
                    raise Exception(f"LibreOffice conversion failed: {result.stderr}")

                actual_pdf_path = temp_excel_path.replace('.xlsx', '.pdf')

                with open(actual_pdf_path, 'rb') as pdf_file:
                    pdf_buffer.write(pdf_file.read())

                if os.path.exists(actual_pdf_path):
                    os.unlink(actual_pdf_path)

            finally:
                if os.path.exists(temp_excel_path):
                    os.unlink(temp_excel_path)
                if os.path.exists(temp_pdf_path):
                    os.unlink(temp_pdf_path)

        except Exception as e:
            raise Exception(f"Error converting Excel to PDF: {e}")

    # def _convert_excel_to_pdf(self, excel_buffer, pdf_buffer):
    #
    #     try:
    #
    #         # Create temporary files
    #         with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_excel:
    #             temp_excel.write(excel_buffer.getvalue())
    #             temp_excel_path = temp_excel.name
    #
    #         with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
    #             temp_pdf_path = temp_pdf.name
    #
    #         try:
    #             # Use LibreOffice headless to convert Excel to PDF
    #             # This preserves the original Excel formatting
    #
    #             result = run([
    #                 'libreoffice',
    #                 '--headless',
    #                 '--convert-to', 'pdf',
    #                 '--outdir', os.path.dirname(temp_pdf_path),
    #                 temp_excel_path
    #             ], capture_output=True, text=True, timeout=30)
    #
    #             if result.returncode != 0:
    #                 raise Exception(f"LibreOffice conversion failed: {result.stderr}")
    #
    #             # LibreOffice creates PDF with same name as Excel but .pdf extension
    #             actual_pdf_path = temp_excel_path.replace('.xlsx', '.pdf')
    #
    #             # Read the generated PDF and write to buffer
    #             with open(actual_pdf_path, 'rb') as pdf_file:
    #                 pdf_buffer.write(pdf_file.read())
    #
    #             # Clean up the generated PDF
    #             if os.path.exists(actual_pdf_path):
    #                 os.unlink(actual_pdf_path)
    #
    #         finally:
    #             # Clean up temporary files
    #             if os.path.exists(temp_excel_path):
    #                 os.unlink(temp_excel_path)
    #             if os.path.exists(temp_pdf_path):
    #                 os.unlink(temp_pdf_path)
    #
    #     except Exception as e:
    #         raise Exception(f"Error converting Excel to PDF: {e}")


    #multi

    # def _convert_multiple_excel_to_pdf(self, excel_buffers, pdf_buffer):
    #
    #     pdf_files_to_merge = []
    #     temp_files_to_cleanup = []
    #
    #     try:
    #         # Convertir cada Excel a PDF
    #         for buffer_index, excel_buffer in enumerate(excel_buffers):
    #             try:
    #                 # Crear archivo temporal Excel
    #                 with tempfile.NamedTemporaryFile(suffix=f'_{buffer_index}.xlsx', delete=False) as temp_excel:
    #                     temp_excel.write(excel_buffer.getvalue())
    #                     temp_excel_path = temp_excel.name
    #                     temp_files_to_cleanup.append(temp_excel_path)
    #
    #                 if buffer_index % 10 == 0 or buffer_index == len(excel_buffers) - 1:
    #                     print(f"Converting file {buffer_index + 1}/{len(excel_buffers)}")
    #
    #                 # Convertir a PDF usando LibreOffice
    #                 result = run([
    #                     'libreoffice',
    #                     '--headless',
    #                     '--convert-to', 'pdf',
    #                     '--outdir', os.path.dirname(temp_excel_path),
    #                     temp_excel_path
    #                 ], capture_output=True, text=True, timeout=120)
    #
    #                 if result.returncode != 0:
    #                     print(f"Warning: LibreOffice conversion failed for file {buffer_index}: {result.stderr}")
    #                     continue
    #
    #                 # LibreOffice crea PDF con el mismo nombre pero extensión .pdf
    #                 pdf_path = temp_excel_path.replace('.xlsx', '.pdf')
    #
    #                 if os.path.exists(pdf_path):
    #                     pdf_files_to_merge.append(pdf_path)
    #                     temp_files_to_cleanup.append(pdf_path)
    #                 else:
    #                     print(f"Warning: PDF not created for file {buffer_index}")
    #
    #             except TimeoutExpired:
    #                 print(f"Warning: Timeout converting file {buffer_index}")
    #                 continue
    #             except Exception as e:
    #                 print(f"Warning: Error converting file {buffer_index}: {e}")
    #                 continue
    #
    #         if not pdf_files_to_merge:
    #             raise Exception("No PDF files were successfully created")
    #
    #         print(f"Successfully converted {len(pdf_files_to_merge)} files. Starting merge with pypdf...")
    #
    #         self._merge_pdfs_with_pypdf_optimized(pdf_files_to_merge, pdf_buffer)
    #
    #         print("✓ PDF merge completed successfully")
    #
    #     except Exception as e:
    #         raise Exception(f"Error converting multiple Excel files to PDF: {e}")
    #
    #     finally:
    #         # Limpiar archivos temporales
    #         self._cleanup_temp_files(temp_files_to_cleanup)

    def _convert_multiple_excel_to_pdf(self, excel_buffers, pdf_buffer):
        pdf_files_to_merge = []
        temp_files_to_cleanup = []

        try:
            # Convertir cada Excel a PDF
            for buffer_index, excel_buffer in enumerate(excel_buffers):
                try:
                    # Crear archivo temporal Excel
                    with tempfile.NamedTemporaryFile(suffix=f'_{buffer_index}.xlsx', delete=False) as temp_excel:
                        temp_excel.write(excel_buffer.getvalue())
                        temp_excel_path = temp_excel.name
                        temp_files_to_cleanup.append(temp_excel_path)

                    if buffer_index % 10 == 0 or buffer_index == len(excel_buffers) - 1:
                        print(f"Converting file {buffer_index + 1}/{len(excel_buffers)}")

                    # ===== APLICAR AJUSTES DE CELDAS =====
                    try:
                        wb = openpyxl.load_workbook(temp_excel_path)

                        # DEFINE LAS COLUMNAS PARA AJUSTAR ALTURA Y ANCHO
                        columns_adjust_height = ['P']
                        columns_adjust_width = ['P']

                        for sheet in wb.worksheets:
                            # PASO 1: Autoajustar ANCHO de columnas específicas
                            for col_letter in columns_adjust_width:
                                max_length = 0

                                for cell in sheet[col_letter]:
                                    # SKIP celdas combinadas
                                    if isinstance(cell, MergedCell):
                                        continue

                                    if cell.value:
                                        # Habilitar wrap_text
                                        if cell.alignment:
                                            cell.alignment = cell.alignment.copy(
                                                wrap_text=True,
                                                vertical='top'
                                            )
                                        else:
                                            cell.alignment = Alignment(
                                                wrap_text=True,
                                                vertical='top'
                                            )

                                        # Calcular longitud máxima
                                        lines = str(cell.value).split('\n')
                                        for line in lines:
                                            max_length = max(max_length, len(line))

                                # Aplicar ancho óptimo
                                if max_length > 0:
                                    optimal_width = min(max_length * 1.2 + 2, 100)
                                    sheet.column_dimensions[col_letter].width = optimal_width

                            # PASO 2: Ajustar ALTURA de filas
                            for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                                max_height = 28  # Altura mínima

                                for cell in row:
                                    # SKIP celdas combinadas
                                    if isinstance(cell, MergedCell):
                                        continue

                                    if cell.value:
                                        # Asegurar que tenga wrap_text
                                        if cell.alignment:
                                            cell.alignment = cell.alignment.copy(
                                                wrap_text=True,
                                                vertical='top'
                                            )
                                        else:
                                            cell.alignment = Alignment(
                                                wrap_text=True,
                                                vertical='top'
                                            )

                                        # Obtener ancho actual de la columna
                                        col_width = sheet.column_dimensions[cell.column_letter].width or 10
                                        chars_per_line = max(int((col_width - 2) / 1.2), 1)

                                        # Calcular líneas necesarias
                                        lines = str(cell.value).split('\n')
                                        total_lines = 0

                                        for line in lines:
                                            if len(line) == 0:
                                                total_lines += 1
                                            else:
                                                total_lines += (len(line) // chars_per_line) + 1

                                        needed_height = total_lines * 5
                                        max_height = max(max_height, needed_height)

                                # Aplicar altura calculada
                                sheet.row_dimensions[row_idx].height = min(max_height, 300)

                        # Guardar los cambios
                        wb.save(temp_excel_path)
                        wb.close()

                    except Exception as e:
                        print(f"Warning: Error adjusting cells for file {buffer_index}: {e}")
                        # Continuar con la conversión aunque falle el ajuste

                    # ===== FIN AJUSTES DE CELDAS =====

                    # Convertir a PDF usando LibreOffice
                    result = run([
                        'libreoffice',
                        '--headless',
                        '--convert-to', 'pdf',
                        '--outdir', os.path.dirname(temp_excel_path),
                        temp_excel_path
                    ], capture_output=True, text=True, timeout=120)

                    if result.returncode != 0:
                        print(f"Warning: LibreOffice conversion failed for file {buffer_index}: {result.stderr}")
                        continue

                    # LibreOffice crea PDF con el mismo nombre pero extensión .pdf
                    pdf_path = temp_excel_path.replace('.xlsx', '.pdf')

                    if os.path.exists(pdf_path):
                        pdf_files_to_merge.append(pdf_path)
                        temp_files_to_cleanup.append(pdf_path)
                    else:
                        print(f"Warning: PDF not created for file {buffer_index}")

                except TimeoutExpired:
                    print(f"Warning: Timeout converting file {buffer_index}")
                    continue
                except Exception as e:
                    print(f"Warning: Error converting file {buffer_index}: {e}")
                    continue

            if not pdf_files_to_merge:
                raise Exception("No PDF files were successfully created")

            print(f"Successfully converted {len(pdf_files_to_merge)} files. Starting merge with pypdf...")

            self._merge_pdfs_with_pypdf_optimized(pdf_files_to_merge, pdf_buffer)

            print("✓ PDF merge completed successfully")

        except Exception as e:
            raise Exception(f"Error converting multiple Excel files to PDF: {e}")

        finally:
            # Limpiar archivos temporales
            self._cleanup_temp_files(temp_files_to_cleanup)

    def _merge_pdfs_with_pypdf_optimized(self, pdf_paths, output_buffer):

        try:
            from pypdf import PdfWriter, PdfReader

            writer = PdfWriter()
            total_pages = 0

            # Procesar archivos en lotes para manejar memoria
            batch_size = 50  # Procesar 50 archivos a la vez

            for batch_start in range(0, len(pdf_paths), batch_size):
                batch_end = min(batch_start + batch_size, len(pdf_paths))
                batch_paths = pdf_paths[batch_start:batch_end]

                for i, pdf_path in enumerate(batch_paths):
                    try:
                        with open(pdf_path, 'rb') as f:
                            reader = PdfReader(f)
                            pages_in_file = len(reader.pages)

                            # Agregar páginas una por una
                            for page_num, page in enumerate(reader.pages):
                                writer.add_page(page)
                                total_pages += 1

                                # Mostrar progreso cada 100 páginas
                                if total_pages % 100 == 0:
                                    print(f"Added {total_pages} pages...")

                            print(f"✓ Added {pages_in_file} pages from file {batch_start + i + 1}")

                    except Exception as e:
                        print(f"Warning: Could not merge {pdf_path}: {e}")
                        continue

                # Liberar memoria del lote actual
                gc.collect()

            print(f"Writing final PDF with {total_pages} total pages...")

            # Escribir el PDF final
            writer.write(output_buffer)
            output_buffer.seek(0)

            print(f"✓ Successfully created PDF with {total_pages} pages")

        except ImportError:
            print("pypdf not installed. Installing...")
            import subprocess
            subprocess.check_call(['pip', 'install', 'pypdf'])

            # Reintentar después de instalar
            from pypdf import PdfWriter, PdfReader
            return self._merge_pdfs_with_pypdf_optimized(pdf_paths, output_buffer)

        except Exception as e:
            raise Exception(f"Error during pypdf merge: {e}")

    def _cleanup_temp_files(self, temp_files_list):
        cleaned = 0
        for temp_file in temp_files_list:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    cleaned += 1
            except Exception as e:
                print(f"Warning: Could not delete temp file {temp_file}: {e}")

        if cleaned > 0:
            print(f"✓ Cleaned up {cleaned} temporary files")



    ####### FIN PDF ######

    @route(["/print/payslips"], type='http', auth='user')
    def get_payroll_report_print(self, list_ids='', **post):
        if not request.env.user.has_group('hr_payroll.group_hr_payroll_user') or not list_ids or re.search("[^0-9|,]", list_ids):
            return request.not_found()

        ids = [int(s) for s in list_ids.split(',')]
        model_payslip = request.env['hr.payslip'].sudo()
        payslips = model_payslip.browse(ids)

        pdf_writer = PdfFileWriter()
        payslip_reports = payslips._get_pdf_reports()

        for report, slips in payslip_reports.items():

            for payslip in slips:
                holidays_data = {}
                holidays_totals = {}

                if payslip.employee_id.contract_id.date_end:
                    holidays_data, holidays_totals = model_payslip.get_holidays_liquidations(
                        payslip.employee_id.contract_id.date_start,
                        payslip.employee_id.contract_id.date_end,
                        payslip.employee_id.id,
                    )

                pdf_content, _ = request.env['ir.actions.report']. \
                    with_context(lang=payslip.employee_id.lang or payslip.env.lang). \
                    sudo(). \
                    _render_qweb_pdf(report, payslip.id, data={
                    'company_id': payslip.company_id,
                    'payslip': payslip,
                    'extra_info': holidays_data,
                    'holidays_totals': holidays_totals
                })
                reader = PdfFileReader(io.BytesIO(pdf_content), strict=False, overwriteWarnings=False)

                for page in range(reader.getNumPages()):
                    pdf_writer.addPage(reader.getPage(page))

        _buffer = io.BytesIO()
        pdf_writer.write(_buffer)
        merged_pdf = _buffer.getvalue()
        _buffer.close()

        if len(payslip_reports) == 1 and len(payslips) == 1 and payslips.struct_id.report_id.print_report_name:
            report_name = safe_eval(payslips.struct_id.report_id.print_report_name, {'object': payslips})
        else:
            report_name = ' - '.join(r.name for r in list(payslip_reports.keys()))
            employees = payslips.employee_id.mapped('name')
            if len(employees) == 1:
                report_name = '%s - %s' % (report_name, employees[0])

        pdfhttpheaders = [
            ('Content-Type', 'application/pdf'),
            ('Content-Length', len(merged_pdf)),
            ('Content-Disposition', content_disposition(report_name + '.pdf'))
        ]

        return request.make_response(merged_pdf, headers=pdfhttpheaders)

    @route(["/get_payroll_warnings"], type="json", auth='user')
    def get_payroll_warning_data(self):
        return request.env['hr.payslip']._get_dashboard_warnings()


    #### aca caclculo de totales y extras

    MONTHS_ES = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr', 5: 'may', 6: 'jun',
        7: 'jul', 8: 'ago', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }

    def _validate_inputs(self, date, employee_id):
        """Valida los parámetros de entrada."""
        if isinstance(date, str):
            try:
                date = datetime.strptime(date, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError("Formato de fecha inválido. Use YYYY-MM-DD.")
        elif isinstance(date, datetime):
            date = date.date()
        elif not isinstance(date, date_module):
            raise ValueError("El parámetro de fecha debe ser válido.")

        if employee_id:
            employee = request.env['hr.employee'].sudo().browse(employee_id)
            if not employee.exists():
                raise ValueError(f"Empleado con ID {employee_id} no existe.")
        return date

    def _get_employee_start_date(self, employee_id):
        """Obtiene la fecha de ingreso del empleado."""
        if not employee_id:
            return None

        employee = request.env['hr.employee'].sudo().browse(employee_id)
        if employee and employee.contract_id.date_start:
            ingreso = employee.contract_id.date_start
            if isinstance(ingreso, str):
                ingreso = datetime.strptime(ingreso, '%Y-%m-%d').date()
            return ingreso
        return None

    def get_vacation_period(self, date_start: date, date_end: date, employee_id: Optional[int]) -> Tuple[
        Union[date, bool], Union[date, bool]]:
        """
        Calcula el período de vacaciones actual basado en las fechas del contrato laboral.

        Args:
            date_start: Fecha de inicio del contrato del empleado
            date_end: Fecha de fin del contrato del empleado (o fecha actual)
            employee_id: ID del empleado

        Returns:
            Tuple con (inicio_periodo_vacaciones, fin_periodo_vacaciones)
            Retorna (False, False) si el empleado tiene menos de un año de antigüedad
        """
        if not employee_id:
            return False, False

        # Verificar si el empleado tiene al menos un año de antigüedad
        if not self._has_minimum_one_year_service(date_start, date_end):
            return False, False

        # Calcular el período de vacaciones actual
        vacation_start, vacation_end = self._get_current_vacation_period(date_start, date_end)

        return vacation_start, vacation_end

    def _has_minimum_one_year_service(self, start_date: date, end_date: date) -> bool:
        """
        Verifica si el empleado tiene al menos un año de servicio.

        Args:
            start_date: Fecha de inicio del contrato
            end_date: Fecha de fin del contrato o fecha actual

        Returns:
            True si tiene al menos un año de antigüedad, False en caso contrario
        """
        try:
            # Calcular el primer aniversario
            first_anniversary = date(start_date.year + 1, start_date.month, start_date.day)
        except ValueError:  # Caso 29 de febrero
            first_anniversary = self._safe_date_creation(start_date.year + 1, start_date.month, start_date.day)

        return end_date >= first_anniversary

    def _get_current_vacation_period(self, contract_start: date, contract_end: date) -> Tuple[date, date]:
        """
        Calcula el período de vacaciones actual en el que se encuentra el empleado.

        Args:
            contract_start: Fecha de inicio del contrato
            contract_end: Fecha de fin del contrato o fecha actual

        Returns:
            Tuple con (inicio_periodo_actual, fin_periodo_actual)
        """
        # Encontrar en qué año de servicio está el empleado
        years_completed = self._calculate_completed_years(contract_start, contract_end)

        # El período de vacaciones actual comienza en el último aniversario
        current_period_start_year = contract_start.year + years_completed

        try:
            vacation_period_start = date(current_period_start_year, contract_start.month, contract_start.day)
        except ValueError:  # Para 29 de febrero
            vacation_period_start = self._safe_date_creation(current_period_start_year, contract_start.month,
                                                             contract_start.day)

        # El período termina en la fecha de fin del contrato o antes del próximo aniversario
        try:
            next_anniversary = date(current_period_start_year + 1, contract_start.month, contract_start.day)
        except ValueError:
            next_anniversary = self._safe_date_creation(current_period_start_year + 1, contract_start.month,
                                                        contract_start.day)

        # El fin del período es el menor entre la fecha de fin del contrato y un día antes del siguiente aniversario
        vacation_period_end = min(contract_end, next_anniversary - timedelta(days=1))

        return vacation_period_start, vacation_period_end

    def _calculate_completed_years(self, start_date: date, current_date: date) -> int:
        """
        Calcula los años completos de servicio del empleado.

        Args:
            start_date: Fecha de inicio del contrato
            current_date: Fecha actual o fin del contrato

        Returns:
            Número entero de años de servicio completados
        """
        years = current_date.year - start_date.year

        # Verificar si ya pasó el aniversario este año
        try:
            anniversary_this_year = date(current_date.year, start_date.month, start_date.day)
        except ValueError:  # Para 29 de febrero
            anniversary_this_year = self._safe_date_creation(current_date.year, start_date.month, start_date.day)

        if current_date < anniversary_this_year:
            years -= 1

        return max(0, years)

    def _safe_date_creation(self, year: int, month: int, day: int) -> date:
        """
        Crea una fecha de forma segura, manejando casos como 29 de febrero.

        Args:
            year: Año
            month: Mes
            day: Día

        Returns:
            Objeto date válido
        """
        try:
            return date(year, month, day)
        except ValueError:
            if month == 2 and day == 29:
                # Si es 29 de febrero y el año no es bisiesto, usar 28 de febrero
                return date(year, month, 28)
            else:
                raise ValueError(f"No se pudo crear fecha válida: {year}-{month}-{day}")

    # Ejemplos de funcionamiento:

    def _get_decimo_tercero_period(self, calculation_date, employee_id):
        """
        Calcula el período válido para décimo tercero sueldo.
        Período legal: 1 de diciembre del año anterior al 30 de noviembre del año actual.
        Pero limitado por la fecha de ingreso del empleado.
        """
        year = calculation_date.year

        # Determinar el período correcto basado en la fecha de cálculo
        if calculation_date.month >= 12:  # Diciembre o después
            # Período actual: dic año actual a nov año siguiente
            legal_start = date_module(year, 12, 1)
            legal_end = date_module(year + 1, 11, 30)
        else:  # Enero a noviembre
            # Período actual: dic año anterior a nov año actual
            legal_start = date_module(year - 1, 12, 1)
            legal_end = date_module(year, 11, 30)

        # Ajustar por fecha de ingreso del empleado
        employee_start = self._get_employee_start_date(employee_id)
        if employee_start:
            # El período efectivo empieza en la fecha de ingreso si es posterior al inicio legal
            effective_start = max(legal_start, employee_start)
            # El período efectivo termina en el último día del mes de cálculo o fin legal
            calculation_month_end = date_module(calculation_date.year, calculation_date.month,
                                                calendar.monthrange(calculation_date.year, calculation_date.month)[1])
            effective_end = min(legal_end, calculation_month_end)
        else:
            effective_start = legal_start
            calculation_month_end = date_module(calculation_date.year, calculation_date.month,
                                                calendar.monthrange(calculation_date.year, calculation_date.month)[1])
            effective_end = min(legal_end, calculation_month_end)

        return effective_start, effective_end

    def _get_decimo_cuarto_period(self, calculation_date, employee_id):
        """
        Calcula el período válido para décimo cuarto sueldo.
        Período legal: 1 de agosto al 31 de julio del año siguiente.
        Pero limitado por la fecha de ingreso del empleado.
        """
        year = calculation_date.year

        # Determinar el período correcto basado en la fecha de cálculo
        if calculation_date.month >= 8:  # Agosto o después
            # Período actual: ago año actual a jul año siguiente
            legal_start = date_module(year, 8, 1)
            legal_end = date_module(year + 1, 7, 31)
        else:  # Enero a julio
            # Período actual: ago año anterior a jul año actual
            legal_start = date_module(year - 1, 8, 1)
            legal_end = date_module(year, 7, 31)

        # Ajustar por fecha de ingreso del empleado
        employee_start = self._get_employee_start_date(employee_id)
        if employee_start:
            # El período efectivo empieza en la fecha de ingreso si es posterior al inicio legal
            effective_start = max(legal_start, employee_start)
            # El período efectivo termina en el último día del mes de cálculo o fin legal
            calculation_month_end = date_module(calculation_date.year, calculation_date.month,
                                                calendar.monthrange(calculation_date.year, calculation_date.month)[1])
            effective_end = min(legal_end, calculation_month_end)
        else:
            effective_start = legal_start
            calculation_month_end = date_module(calculation_date.year, calculation_date.month,
                                                calendar.monthrange(calculation_date.year, calculation_date.month)[1])
            effective_end = min(legal_end, calculation_month_end)

        return effective_start, effective_end
    def _get_earliest_date(self, date_start: date, date_end: date, employee_id: Optional[int]) -> Union[
        date, bool]:

        if not employee_id:
            return False

        # Obtener todas las fechas de inicio
        vacation_start, _ = self.get_vacation_period(date_start, date_end, employee_id)
        dt_start, _ = self._get_decimo_tercero_period(date_end, employee_id)
        dc_start, _ = self._get_decimo_cuarto_period(date_end, employee_id)

        # Filtrar fechas válidas usando comprensión de lista
        valid_dates = [
            start_date for start_date in [vacation_start, dt_start, dc_start]
            if start_date is not False
        ]

        return min(valid_dates) if valid_dates else False

    def _is_month_in_period(self, month_date: date, period_start: Union[date, bool],
                           period_end: Union[date, bool]) -> bool:

        if period_start is False or period_end is False:
            return False

        year = month_date.year
        month = month_date.month

        month_start = date(year, month, 1)

        # Último día del mes
        days_in_month = monthrange(year, month)[1]
        month_end = date(year, month, days_in_month)

        return period_start <= month_end and period_end >= month_start

    def _build_payslip_domain(self, date_from, date_to, employee_id):
        domain = [
            ('date_from', '>=', date_from),
            ('date_to', '<=', date_to),
            ('state', 'in', ['done', 'paid']),
            ('struct_id.name', '=', 'Rol de Pagos'),
        ]
        if employee_id:
            domain.append(('employee_id', '=', employee_id))
        return domain

    def get_holidays_liquidations(self, date_start=False, date_end=False, employee_id=False, payslip=False):
        """
        Genera un reporte de nóminas con neto, neto/24 y décimos para el período de vacaciones.
        Considera correctamente los períodos de décimos basándose en la fecha de ingreso del empleado.

        :param date: Fecha límite (str o date, por defecto fecha actual).
        :param employee_id: ID del empleado (opcional).
        :return: Lista de diccionarios con datos de nóminas por mes.
        """
        # Establecer valores por defecto
        date = date_end or datetime.now().date()
        date = self._validate_inputs(date_end, employee_id)

        vacation_start, vacation_end = self.get_vacation_period(date_start, date_end, employee_id)
        dt_period_start, dt_period_end = self._get_decimo_tercero_period(date_end, employee_id)
        dc_period_start, dc_period_end = self._get_decimo_cuarto_period(date_end, employee_id)

        date_from = self._get_earliest_date(date_start, date_end, employee_id)
        date_to = date

        # Construir dominio y buscar nóminas
        domain = self._build_payslip_domain(date_from, date_to, employee_id)
        payslips = request.env['hr.payslip'].sudo().search(domain)
        payslips = payslips + payslip

        # Generar meses en el rango
        all_months = []
        totals = {}
        current_date = date_from.replace(day=1)
        if vacation_start and vacation_end:
            days_holidays = self.calculate_days_worked(vacation_start, vacation_end)
        else:
            days_holidays = 0

        while current_date <= date_to:
            all_months.append((current_date.year, current_date.month))
            current_date += relativedelta(months=1)


        # Obtener empleados
        if employee_id:
            employee_ids = [employee_id]
        else:
            employee_ids = list(set([p.employee_id.id for p in payslips]))
            if not employee_ids and employee_id:
                employee_ids = [employee_id]

        result = []
        payslips_by_month = {}

        # Agrupar nóminas por mes y empleado
        for payslip in payslips:
            emp_id = payslip.employee_id.id
            payslip_date = payslip.date_from
            month_key = f"{emp_id}-{payslip_date.year}-{payslip_date.month}"
            if month_key not in payslips_by_month or payslip.date_from > payslips_by_month[month_key].date_from:
                payslips_by_month[month_key] = payslip

        # Generar reporte
        for emp_id in employee_ids:

            totals = {
                "neto": 0.0,
                "holidays": 0.0,
                "dt": 0.0,
                "dc": 0.0,
                "days": 0
            }

            for year, month in all_months:
                month_str = self.MONTHS_ES[month]
                year_str = str(year)[-2:]
                mes_formato = f"{month_str}-{year_str}"
                month_key = f"{emp_id}-{year}-{month}"

                # Crear fecha del primer día del mes para comparaciones
                month_date = date_module(year, month, 1)

                # Determinar si el mes está en los períodos (usando los períodos ajustados)
                is_vacation_period = self._is_month_in_period(month_date, vacation_start, vacation_end)
                is_dec_tercero_period = self._is_month_in_period(month_date, dt_period_start, dt_period_end)
                is_dec_cuarto_period = self._is_month_in_period(month_date, dc_period_start, dc_period_end)

                payslip_dict = {
                    "month": mes_formato,
                    "neto": "N/A",
                    "days": "" if not is_vacation_period or not is_dec_tercero_period or not is_dec_cuarto_period else "N/A",
                    "holidays": "" if not is_vacation_period else "N/A",
                    "dt": "" if not is_dec_tercero_period else "N/A",
                    "dc": "" if not is_dec_cuarto_period else "N/A",
                    "has_payslip": False,
                    "totals": {
                        "neto": 0.0,
                        "holidays": 0.0,
                        "dt": 0.0,
                        "dc": 0.0,
                        "days": 0
                    },
                    "period_info": {
                        "vacation_period": is_vacation_period,
                        "dt_period": is_dec_tercero_period,
                        "dc_period": is_dec_cuarto_period,
                        "month_date": f"{year}-{month:02d}-01"
                    }
                }

                if month_key in payslips_by_month:
                    payslip = payslips_by_month[month_key]
                    payslip_dict["has_payslip"] = True

                    for line in payslip.line_ids:
                        if line.code == 'LIQAPAG':
                            neto_amount = round(line.amount, 2)
                            payslip_dict["neto"] = neto_amount
                            totals["neto"] += neto_amount

                            if is_vacation_period:
                                days = days_holidays[str(month_date)]
                                holidays_amount = round(((line.amount * int(days)) / 30)/ 24, 2) if line.amount else 0
                                payslip_dict["holidays"] = holidays_amount
                                totals["holidays"] += holidays_amount

                        elif line.code == 'DECTER' and is_dec_tercero_period:
                            dt_amount = round(line.amount, 2)
                            payslip_dict["dt"] = dt_amount
                            totals["dt"] += dt_amount

                        elif line.code == 'DECCUAR' and is_dec_cuarto_period:
                            dc_amount = round(line.amount, 2)
                            payslip_dict["dc"] = dc_amount
                            totals["dc"] += dc_amount

                        if line.code == 'DYSMES':
                            days_amount = int(line.amount)
                            payslip_dict["days"] = days_amount
                            totals["days"] += days_amount

                result.append(payslip_dict)

        return result, totals

    def calculate_days_worked(self, date_start, date_end):
        """
        Calcula los días trabajados por mes en el período específico del contrato.

        Args:
            date_start (str or date): Fecha de inicio del contrato
            date_end (str or date): Fecha de fin del contrato

        Returns:
            dict: Diccionario con formato {'2024-01-01': dias, '2024-02-01': dias, ...}
        """
        # Convertir strings a objetos date si es necesario
        if isinstance(date_start, str):
            contract_start = date.fromisoformat(date_start)
        else:
            contract_start = date_start

        if isinstance(date_end, str):
            contract_end = date.fromisoformat(date_end)
        else:
            contract_end = date_end

        result = {}

        # Obtener el rango de años y meses a procesar

        start_year = contract_start.year
        start_month = contract_start.month
        end_year = contract_end.year
        end_month = contract_end.month

        # Iterar por todos los meses en el rango del contrato
        current_year = start_year
        current_month = start_month

        while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
            # Calcular inicio y fin del mes actual
            month_start = date(current_year, current_month, 1)

            # Calcular el último día del mes
            if current_month == 12:
                next_month_start = date(current_year + 1, 1, 1)
            else:
                next_month_start = date(current_year, current_month + 1, 1)
            month_end = next_month_start - timedelta(days=1)

            # Verificar si el contrato abarca este mes
            if contract_start <= month_end and contract_end >= month_start:
                if contract_start <= month_start and contract_end >= month_end:
                    # Mes completo trabajado - usar días reales del mes
                    days_worked = (month_end - month_start).days + 1
                else:
                    # Mes parcialmente trabajado
                    work_start = max(contract_start, month_start)
                    work_end = min(contract_end, month_end)

                    # Calcular días reales trabajados
                    days_worked = (work_end - work_start).days + 1

                # Agregar al resultado con formato YYYY-MM-01
                month_key = f"{current_year}-{current_month:02d}-01"
                result[month_key] = int(days_worked)

            # Avanzar al siguiente mes
            if current_month == 12:
                current_year += 1
                current_month = 1
            else:
                current_month += 1

        # Ajustar el último mes a 30 días siempre
        if result:
            # Obtener la última clave (último mes)
            last_month_key = max(result.keys())
            result[last_month_key] = 30

        return result

    def get_months_without_payslips(self, date=False, employee_id=False):
        """
        Retorna solo los meses que no tienen payslips (aparecerán con string vacío).
        """
        all_data = self.get_holidays_liquidations(date, employee_id)
        months_without_data = [record for record in all_data if not record.get("has_payslip", False)]

        return {
            "total_months": len(all_data),
            "months_without_payslips": len(months_without_data),
            "missing_months": [{"mes": r["month"], "employee_id": employee_id} for r in months_without_data]
        }


    meses_espanol = {
        "January": "Enero",
        "February": "Febrero",
        "March": "Marzo",
        "April": "Abril",
        "May": "Mayo",
        "June": "Junio",
        "July": "Julio",
        "August": "Agosto",
        "September": "Septiembre",
        "October": "Octubre",
        "November": "Noviembre",
        "December": "Diciembre",
    }

    ## no Borrar, funcion para generar contexto, puede ser usada para generar pdf a partir de plantilla xml
    # @http.route('/reporte_asistencias/download_pdf/<int:record_id>', type='http', auth='public')
    # def download_pdf(self, record_id, **kwargs):
    #     # Validar permisos y registro
    #     if not request.env.user.has_group('hr_payroll.group_hr_payroll_user'):
    #         return request.not_found()
    #
    #     record = request.env['report.attendance.general'].sudo().browse(record_id)
    #     if not record.exists():
    #         return request.not_found()
    #
    #     # Validar empleados
    #     if not record.employee_ids:
    #         return request.make_response(
    #             'No employees found for this report.',
    #             headers=[('Content-Type', 'text/plain')]
    #         )
    #
    #     # Cache system parameter
    #     type_of_resource = request.env['ir.config_parameter'].sudo().get_param('hr_payroll.mode_of_attendance')
    #     work_entry_env = request.env['hr.work.entry'].sudo()
    #
    #     # Definir rango de fechas
    #     date_from = datetime.combine(record.date_from, time.min)
    #     date_to = datetime.combine(record.date_to, time.max)
    #     date_range = [
    #         date_from + timedelta(days=x)
    #         for x in range((record.date_to - record.date_from).days + 1)
    #     ]
    #     date_utc_ranges = [
    #         (
    #             record.convert_to_utc(datetime.combine(date, time.min) + timedelta(minutes=1)),
    #             record.convert_to_utc(datetime.combine(date, time.max) - timedelta(minutes=1))
    #         )
    #         for date in date_range
    #     ]
    #
    #     # Obtener entradas de trabajo
    #     work_entries = work_entry_env.search([
    #         ('employee_id', 'in', record.employee_ids.ids),
    #         ('date_start', '>=', date_utc_ranges[0][0]),
    #         ('date_stop', '<=', date_utc_ranges[-1][1]),
    #     ])
    #     work_entries_dict = defaultdict(list)
    #     for entry in work_entries:
    #         date_str = record.convertir_a_hora_ecuador(entry.date_start).date().isoformat()
    #         work_entries_dict[(entry.employee_id.id, date_str)].append(entry)
    #
    #     # Precargar datos de festivos y calendario
    #     holidays_dict = record._prefetch_holidays(record.employee_ids.ids, date_utc_ranges)
    #     calendar_dict = record._prefetch_calendar(record.employee_ids.ids, date_range, type_of_resource)
    #
    #     # Generar datos del reporte
    #     report_data = []
    #     for employee in record.employee_ids:
    #         employee_data = {
    #             'employee_name': employee.name,
    #             'company_name': employee.company_id.name or '',
    #             'department_name': employee.department_id.name or 'Sin Departamento',
    #             'identification_id': employee.identification_id or '',
    #             'entries': [],
    #             'totals': {
    #                 'total_days_work': 0,
    #                 'total_hours_mount': 0,
    #                 'total_hours_normal': 0,
    #                 'total_hours_delays': 0,
    #                 'total_hours_nocturne': 0,
    #                 'total_hours_supplementary': 0,
    #                 'total_hours_extraordinary': 0,
    #                 'total_hours_debit': 0,
    #             }
    #         }
    #
    #         for current_date in date_range:
    #             date_str = current_date.isoformat()
    #             work_entry_date = work_entries_dict.get((employee.id, date_str), [])
    #             work_entry = record.get_attendances_with_incosistencies(
    #                 record.convert_to_utc(datetime.combine(current_date, time.min)),
    #                 record.convert_to_utc(datetime.combine(current_date, time.max)),
    #                 employee
    #             )
    #
    #             holidays = holidays_dict.get((employee.id, date_str), {'national': False, 'province': False})
    #             calendar = calendar_dict.get((employee.id, date_str), False)
    #
    #             # Calcular datos de la entrada
    #             entry_data = {
    #                 'date': current_date.strftime('%d/%m/%Y'),
    #                 'day_name': TRANSLATE_DAYS.get(current_date.strftime('%A'), ''),
    #                 'is_leave': record._check_is_leave(work_entry, 0),
    #                 'total_days_work': 0,
    #                 'hours_delays': 0,
    #                 'hours_nocturne': 0,
    #                 'hours_supplementary': 0,
    #                 'hours_extraordinary': 0,
    #                 'hours_debit': 0,
    #                 'reference': record.get_references(work_entry),
    #                 'observation': '',
    #             }
    #
    #             # Calcular días trabajados
    #             total_days_work, employee_data['totals']['total_days_work'] = record.calculate_total_hour(
    #                 work_entry_date, employee_data['totals']['total_days_work']
    #             )
    #             entry_data['total_days_work'] = total_days_work
    #
    #             # Calcular tipos de horas
    #             hour_types = [
    #                 ('hours_delays', 'total_hours_delays', 'hr_payroll.hr_work_entry_type_delays'),
    #                 ('hours_nocturne', 'total_hours_nocturne', 'hr_payroll.hr_work_entry_type_nocturne'),
    #                 ('hours_supplementary', 'total_hours_supplementary', 'hr_payroll.hr_work_entry_type_sumplementary'),
    #                 ('hours_extraordinary', 'total_hours_extraordinary', 'hr_payroll.hr_work_entry_type_extraordinary'),
    #                 ('hours_debit', 'total_hours_debit', 'hr_payroll.hr_work_entry_type_leaves'),
    #             ]
    #             for field, total_key, entry_type in hour_types:
    #                 value, employee_data['totals'][total_key] = record._calculate_total_hours(
    #                     work_entry_date, employee_data['totals'][total_key], entry_type
    #                 )
    #                 entry_data[field] = value
    #
    #             # Verificar asistencia
    #             observation, employee_data['totals']['total_days_work'], employee_data['totals'][
    #                 'total_hours_mount'] = record._check_attendance_exist(
    #                 work_entry, work_entry_date, employee.id, current_date,
    #                 datetime.combine(current_date, time.min), datetime.combine(current_date, time.max),
    #                 employee_data['totals']['total_days_work'], employee_data['totals']['total_hours_mount'],
    #                 holidays, calendar
    #             )
    #             entry_data['observation'] = observation
    #
    #             employee_data['entries'].append(entry_data)
    #
    #         report_data.append(employee_data)
    #
    #     # Generar el PDF
    #     if not report_data:
    #         return request.make_response(
    #             'No hay datos para generar el reporte.',
    #             headers=[('Content-Type', 'text/plain')]
    #         )
    #
    #
    #     pdf, _ = request.env['ir.actions.report'].sudo()._render_qweb_pdf(
    #         'hr_payroll.hr_attendance_report_pdff',
    #         [record.id], data={
    #             'report_data': report_data,
    #             'date_from': record.date_from,
    #             'date_to': record.date_to,
    #             'month_name': record.date_from.strftime('%B').upper()
    #         }
    #     )
    #
    #     # Definir el nombre del reporte
    #     report_name = f'Reporte_Asistencias_{record.date_from.strftime("%Y%m%d")}_{record.date_to.strftime("%Y%m%d")}'
    #     if len(record.employee_ids) == 1:
    #         report_name = f'{report_name}_{record.employee_ids[0].name}'
    #
    #     # Devolver el PDF
    #     pdfhttpheaders = [
    #         ('Content-Type', 'application/pdf'),
    #         ('Content-Length', len(pdf)),
    #         ('Content-Disposition', content_disposition(report_name + '.pdf')),
    #     ]
    #     return request.make_response(pdf, headers=pdfhttpheaders)

    @http.route('/reporte_asistencias/download_xlsx/<int:record_id>', type='http', auth='public')
    def download_xlsx_of_report(self, record_id, **kwargs):
        # Early validation of record
        record = request.env['report.attendance.general'].sudo().browse(record_id)
        model_import = request.env['hr.attendance.import'].sudo()
        if not record.exists():
            return request.not_found()

        # Use employee_ids directly from record
        employees = record.employee_ids
        if not employees:
            return request.make_response(
                b"No employees found for this report.",
                headers=[('Content-Type', 'text/plain')]
            )

        # Get month name in Spanish
        name_mount = record.date_from.strftime("%B")
        name_mount = MESES_ESPANOL.get(name_mount, name_mount)

        # Prefetch holidays and calendar data
        date_range = [
            record.date_from + timedelta(days=x)
            for x in range((record.date_to - record.date_from).days + 1)
        ]
        date_utc_ranges = [
            (
                record.convert_to_utc(datetime.combine(date, time.min) + timedelta(minutes=1)),
                record.convert_to_utc(datetime.combine(date, time.max) - timedelta(minutes=1))
            )
            for date in date_range
        ]

        holidays_dict = record._prefetch_holidays(employees, date_utc_ranges)
        calendar_dict = model_import.get_range_resource_calendar_massive(employees.ids, date_range,
                                                                         self.type_of_resource)
        holidays_employee_dict = record._prefetch_holidays_employee(employees.ids, date_utc_ranges)
        holidays_employee_permits_dict_all = record._prefetch_holidays_employee_permits_all(employees.ids,
                                                                                            date_utc_ranges)
        schedules_names_by_employee = record.prefetch_calendar_names(employees.ids, record.date_from, record.date_to)

        # Prefetch employee IDs
        employee_ids = employees.ids

        # Define date range in UTC
        date_from = record.convert_to_utc(datetime.combine(record.date_from, time.min))
        date_to = record.convert_to_utc(datetime.combine(record.date_to, time.max))

        work_attendances = self.process_mega_dataset(employee_ids, date_range)

        work_entries = request.env['hr.work.entry'].sudo().search([
            ('employee_id', 'in', employee_ids),
            ('date_start', '>=', date_from),
            ('date_stop', '<=', date_to),
        ])


        # If 20 employees or less, generate single Excel file
        if len(employees) <= 20:
            xlsx_buffer = io.BytesIO()
            try:
                workbook = xlsxwriter.Workbook(xlsx_buffer, {'in_memory': True})
                self._generate_workbook_for_group(
                    workbook,
                    employees,
                    record,
                    name_mount,
                    holidays_dict,
                    calendar_dict,
                    holidays_employee_dict,
                    holidays_employee_permits_dict_all,
                    schedules_names_by_employee,
                    date_range=date_range,
                    work_entries=work_entries,
                    model_import=model_import,
                    work_attendances=work_attendances
                )
                workbook.close()
                xlsx_buffer.seek(0)

                response = request.make_response(
                    xlsx_buffer.getvalue(),
                    headers=[
                        ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                        ('Content-Disposition', content_disposition(f'reporte_asistencias_{name_mount}.xlsx'))
                    ]
                )
                return response
            finally:
                xlsx_buffer.close()

        # If more than 20 employees, generate ZIP with multiple Excel files
        else:
            # Batch employees for memory efficiency
            employees_per_file = 20
            employee_groups = (
                employees[i:i + employees_per_file]
                for i in range(0, len(employees), employees_per_file)
            )

            # Stream ZIP file
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for group_index, employee_group in enumerate(employee_groups, start=1):
                    if not employee_group:
                        continue
                    xlsx_buffer = io.BytesIO()
                    try:
                        workbook = xlsxwriter.Workbook(xlsx_buffer, {'in_memory': True})
                        self._generate_workbook_for_group(
                            workbook,
                            employee_group,
                            record,
                            name_mount,
                            holidays_dict,
                            calendar_dict,
                            holidays_employee_dict,
                            holidays_employee_permits_dict_all,
                            schedules_names_by_employee,
                            date_range=date_range,
                            work_entries=work_entries,
                            model_import=model_import,
                            work_attendances=work_attendances
                        )
                        workbook.close()
                        xlsx_buffer.seek(0)
                        zip_file.writestr(
                            f"reporte_asistencias_{name_mount}_grupo_{group_index}.xlsx",
                            xlsx_buffer.getvalue()
                        )
                    finally:
                        xlsx_buffer.close()

            zip_buffer.seek(0)
            response = request.make_response(
                zip_buffer.getvalue(),
                headers=[
                    ('Content-Type', 'application/zip'),
                    ('Content-Disposition', content_disposition(f'reporte_asistencias_{name_mount}.zip'))
                ]
            )
            zip_buffer.close()
            return response

    def _generate_workbook_for_group(
            self,
            workbook,
            employees,
            record,
            name_mount,
            holidays_dict,
            calendar_dict,
            holidays_employee_dict,
            holidays_employee_permits_dict_all,
            schedules_names_by_employee,
            date_range,
            work_entries,
            model_import,
            work_attendances
    ):
        # Define formats once
        border_format = workbook.add_format({"border": 1})
        header_format_title = workbook.add_format({"bold": True, "align": "center", "valign": "vcenter", "font_size": 18})
        header_format2 = workbook.add_format({"bold": True, "align": "left", "valign": "vcenter", "font_size": 16})
        header_format3 = workbook.add_format({"border": 1,"bold": True, "align": "left", "valign": "vcenter", "font_size": 14})
        footer_format = workbook.add_format({"bold": True, "align": "left", "valign": "vcenter", "font_size": 14})
        font_general = workbook.add_format({"border": 1, "bold": False, "align": "left", "valign": "vcenter", "font_size": 14})

        # Build work entries dictionary efficiently
        work_entries_dict = defaultdict(list)

        for entry in work_entries:
            date_str = record.convertir_a_hora_ecuador(entry.date_start).date().isoformat()
            work_entries_dict[(entry.employee_id.id, date_str)].append(entry)

        for employee in employees:
            sheet_name = self._generate_unique_sheet_name(workbook, employee.name)
            worksheet = workbook.add_worksheet(sheet_name)

            # Write headers
            for col_num, header in enumerate(HEADERS):
                worksheet.write(5, col_num, header, header_format3)

            # Write employee header
            self._write_employee_header(
                worksheet,
                employee,
                record,
                name_mount,
                len(HEADERS),
                header_format_title,
                header_format2,
                font_general
            )

            page_setup = {
                'orientation': request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_orientation',
                                                                                   'landscape'),
                'fit_to_width': int(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_fit_to_width', '1')),
                'fit_to_height': int(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_fit_to_height', '0')),
                'margin_left': float(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_margin_left', '0.1')),
                # 0.1 pulgadas (~2.5mm)
                'margin_right': float(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_margin_right', '0.1')),
                'margin_top': float(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_margin_top', '0.1')),
                'margin_bottom': float(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_margin_bottom', '0.1')),
                'center_horizontally': request.env['ir.config_parameter'].sudo().get_param(
                    'hr_payroll.page_center_horizontally', 'True') == 'True',
                'paper_size': int(
                    request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_paper_size', '9')),
                # A4 por defecto
                'use_landscape': request.env['ir.config_parameter'].sudo().get_param('hr_payroll.page_orientation',
                                                                                     'landscape') == 'landscape'
            }

            if page_setup['use_landscape']:
                worksheet.set_landscape()

            worksheet.set_paper(page_setup['paper_size'])
            worksheet.fit_to_pages(page_setup['fit_to_width'], page_setup['fit_to_height'])

            # Configurar márgenes (en pulgadas)
            worksheet.set_margins(
                left=page_setup['margin_left'],
                right=page_setup['margin_right'],
                top=page_setup['margin_top'],
                bottom=page_setup['margin_bottom']
            )

            totals = self._initialize_totals()
            row = 6

            employee_attendances = work_attendances.get(employee.id, {})

            for current_date in date_range:
                date_str = current_date.isoformat()
                work_entry_date = work_entries_dict.get((employee.id, date_str), [])
                # work_entry = record.get_attendances_with_incosistencies(
                #     record.convert_to_utc(datetime.combine(current_date, time.min)),
                #     record.convert_to_utc(datetime.combine(current_date, time.max)),
                #     employee
                # )
                work_entry = employee_attendances.get(current_date, [])
                totals = self._write_employee_data(
                    workbook,
                    worksheet,
                    row,
                    current_date,
                    TRANSLATE_DAYS,
                    work_entry,
                    work_entry_date,
                    border_format,
                    font_general,
                    totals,
                    employee,
                    holidays_dict,
                    holidays_employee_dict,
                    holidays_employee_permits_dict_all,
                    calendar_dict,
                    record,
                    model_import=model_import
                )
                row += 1

            self._write_totals(
                workbook,
                worksheet,
                row,
                totals,
                border_format,
                font_general,
                footer_format,
                record,
                employee,
                schedules_names_by_employee
            )

        return workbook

    def _generate_unique_sheet_name(self, workbook, base_name):
        sheet_name = base_name[:27]
        counter = 1
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name[:27]} ({counter})"
            counter += 1
        return sheet_name

    def _write_employee_header(self, worksheet, employee, record, name_mount, col_count, header_format_title, header_format2, font_general):
        worksheet.merge_range(0, 0, 0, col_count - 1,
                              f"{(employee.company_id.name or '').upper()}  Sucursal: {(employee.department_id.name or 'SIN DEPARTAMENTO').upper()}",
                              header_format_title)
        worksheet.merge_range(1, 0, 1, col_count - 1, "Reporte General de Asistencia del Empleado", header_format_title)
        worksheet.merge_range(2, 0, 2, 1, "Nombre:", header_format2)
        worksheet.merge_range(3, 0, 3, 1, "Descripción:", header_format2)
        worksheet.merge_range(4, 0, 4, 1, "Cédula:", header_format2)
        worksheet.merge_range(2, 2, 2, 7, employee.name.upper(), header_format2)
        worksheet.merge_range(3, 2, 3, 8,
                              f"CALCULO DE HORAS DEL MES {name_mount.upper()} {record.date_from.year}",
                              header_format2)
        worksheet.merge_range(4, 2, 4, 7, employee.identification_id or '', header_format2)
        worksheet.merge_range(3, 9, 3, 11, f"Desde: {record.date_from.strftime('%d/%m/%Y')}", header_format2)
        worksheet.merge_range(3, 12, 3, 14, f"Hasta: {record.date_to.strftime('%d/%m/%Y')}", header_format2)

    def _initialize_totals(self):
        return {
            'total_days_work': 0,
            'total_hours_mount': 0,
            'total_hours_delays': 0,
            'total_hours_nocturne': 0,
            'total_hours_supplementary': 0,
            'total_hours_extraordinary': 0,
            'total_hours_debit': 0,
        }

    def _write_employee_data(
            self,
            workbook,
            worksheet,
            row,
            current_date,
            translate,
            work_entry,
            work_entry_date,
            border_format,
            font_general,
            totals,
            employee,
            holidays_dict,
            holidays_employee_dict,
            holidays_employee_permits_dict_all,
            calendar_dict,
            record,
            model_import
    ):
        current_date_start = datetime.combine(current_date, time.min)
        current_date_stop = datetime.combine(current_date, time.max)

        # Write basic data
        worksheet.write(row, 0, current_date.strftime("%d/%m/%Y"), font_general)
        worksheet.set_column(0, 0, self.get_column_width([current_date.strftime("%d/%m/%Y")]))

        worksheet.write(row, 1, translate.get(current_date.strftime("%A"), ''), font_general)
        for i in range(6):
            worksheet.write(row, 2 + i, record._check_is_leave(work_entry, i), font_general)

        # Calculate hours
        total_days_work, totals['total_days_work'] = record.calculate_total_hour(
            work_entry_date, totals['total_days_work']
        )
        worksheet.write(row, 8, total_days_work, font_general)

        # Calculate various hour types
        hour_types = [
            ('total_hours_delays', 'hr_payroll.hr_work_entry_type_delays'),
            ('total_hours_nocturne', 'hr_payroll.hr_work_entry_type_nocturne'),
            ('total_hours_supplementary', 'hr_payroll.hr_work_entry_type_sumplementary'),
            ('total_hours_extraordinary', 'hr_payroll.hr_work_entry_type_extraordinary'),
            ('total_hours_debit', 'hr_payroll.hr_work_entry_type_leaves'),
        ]
        for idx, (key, entry_type) in enumerate(hour_types, start=9):
            value, totals[key] = record._calculate_total_hours(work_entry_date, totals[key], entry_type)
            worksheet.write(row, idx, value, font_general)

        font_general_min = workbook.add_format(
            {"border": 1, "bold": False, "align": "left", "valign": "vcenter", "font_size": 11})

        worksheet.write(row, 14, record.get_references(work_entry), font_general_min)
        worksheet.set_column(14, 14, self.get_column_width([record.get_references(work_entry) + "       "]))
        # Check attendance
        date_str = current_date.isoformat()
        holidays = holidays_dict.get((employee, current_date.strftime('%Y-%m-%d')), {'national': False, 'province': False})

        holidays_employee = holidays_employee_dict.get((employee.id, current_date.strftime('%Y-%m-%d')), None)
        holidays_employee_permits_all = holidays_employee_permits_dict_all.get((employee.id, current_date.strftime('%Y-%m-%d')), None)
        # calendar = calendar_dict.get((employee.id, date_str), False)
        calendar = calendar_dict.get(employee.id, {}).get(current_date, {}).get('history', {})

        ranges_contracts = calendar.get('ranges', [])
        max_hours_for_schedule = calendar.get('max_hours', 0)
        is_especial_turn = calendar.get('is_special_shift', False)
        is_extraordinary = calendar.get('is_extraordinary', False)

        # 4. Manejo de turnos especiales
        if is_especial_turn:
            next_day = current_date + timedelta(days=1)
            next_ranges = calendar_dict.get(employee.id, {}).get(next_day, {}).get('history', {})
            ranges_contracts.extend(next_ranges.get('ranges', []))
            ranges_contracts = model_import.filtrar_turno_especial(ranges_contracts)

        # 5. Si no hay rangos, intentar con modo 'employee' o 'departament'
        if not ranges_contracts:
            type_of_resource = request.env['ir.config_parameter'].sudo().get_param(
                'hr_payroll.mode_of_attendance')
            if type_of_resource == 'employee':
                # Usar rangos pre-cargados para 'employee'
                emp_ranges = calendar_dict.get(employee.id, {}).get(date_str, {}).get('employee', {})
                ranges_contracts = emp_ranges.get('ranges', [])
                max_hours_for_schedule = emp_ranges.get('max_hours', 0)
                is_especial_turn = emp_ranges.get('is_special_shift', False)
                is_extraordinary = emp_ranges.get('is_extraordinary', False)

                if is_especial_turn:
                    next_day = current_date + timedelta(days=1)
                    next_ranges = calendar_dict.get(employee.id, {}).get(next_day, {}).get('employee', {})
                    ranges_contracts.extend(next_ranges.get('ranges', []))
                    ranges_contracts = model_import.filtrar_turno_especial(ranges_contracts)
            elif type_of_resource == 'departament':
                # Mantener lógica existente para departamentos
                (
                    ranges_contracts,
                    max_hours_for_schedule,
                    is_especial_turn
                ) = model_import.get_range_resource_calendar_for_departament(
                    employee.id, date_str, date_str
                )
        observation, totals['total_days_work'], totals['total_hours_mount'] = record._check_attendance_exist(
            work_entry,
            work_entry_date,
            employee.id,
            current_date,
            current_date_start,
            current_date_stop,
            totals['total_days_work'],
            totals['total_hours_mount'],
            holidays,
            holidays_employee,
            holidays_employee_permits_all,
            calendar
        )

        worksheet.set_column(14, 14, 20)
        worksheet.set_column(15, 15, 20)
        worksheet.write(row, 15, observation, font_general)
        worksheet.set_column(15, 15, self.get_column_width([observation]))


        return totals

    def _write_totals(
            self,
            workbook,
            worksheet,
            row,
            totals,
            border_format,
            font_general,
            footer_format,
            record,
            employee,
            schedules_names_by_employee
    ):

        worksheet.merge_range(row + 1, 0, row + 1, 1, "Total días laborables:", footer_format)
        worksheet.write(row + 1, 2, record._format_seconds_to_hhmm(totals['total_days_work']), font_general)
        worksheet.merge_range(row + 2, 0, row + 2, 1, "Total horas mes:", footer_format)
        worksheet.write(row + 2, 2, record._format_seconds_to_hhmm(totals['total_hours_mount']), font_general)
        worksheet.merge_range(row + 1, 4, row + 1, 6, "Total horas atrasos:", footer_format)
        worksheet.write(row + 1, 7, record.convert_in_hour_format(totals['total_hours_delays']), font_general)
        worksheet.merge_range(row + 2, 4, row + 2, 6, "Total horas debe:", footer_format)
        worksheet.write(row + 2, 7, record.convert_in_hour_format(totals['total_hours_debit']), font_general)
        worksheet.merge_range(row + 1, 9, row + 1, 10, "Total horas 25%:", footer_format)
        worksheet.write(row + 1, 11, record.convert_in_hour_format(totals['total_hours_nocturne']), font_general)
        worksheet.merge_range(row + 2, 9, row + 2, 10, "Total horas 50%:", footer_format)
        worksheet.write(row + 2, 11, record.convert_in_hour_format(totals['total_hours_supplementary']), font_general)
        worksheet.merge_range(row + 3, 9, row + 3, 10, "Total horas 100%:", footer_format)
        worksheet.write(row + 3, 11, record.convert_in_hour_format(totals['total_hours_extraordinary']), font_general)
        worksheet.merge_range(row + 1, 13, row + 1, 16, "Calculo Horas Extra:", footer_format)

        supplementary_hours, extraordinary_hours = record.calculate_total_hours_extra(
            totals['total_hours_debit'], totals['total_hours_supplementary'], totals['total_hours_extraordinary']
        )

        worksheet.write(row + 2, 13, "25%:", footer_format)
        worksheet.write(row + 3, 13, record.convert_in_hour_format(totals['total_hours_nocturne']), font_general)

        worksheet.write(row + 2, 14, "50%:", footer_format)
        worksheet.write(row + 3, 14, supplementary_hours, font_general)

        worksheet.write(row + 2, 15, "100%:", footer_format)
        worksheet.write(row + 3, 15, extraordinary_hours, font_general)

        # worksheet.write(row + 4, 12, "Horarios:", footer_format)
        #
        # calendar_names = schedules_names_by_employee.get(employee.id, [])
        # worksheet.set_column(13, 13, 25)

        # if calendar_names:
        #     for i, calendar_name in enumerate(calendar_names):
        #         worksheet.write(row + 4 + i, 13, calendar_name)
        # else:
        #     worksheet.write(row + 4, 13, 'Sin calendario')



    def convert_to_utc(self, hour_ecuador):
        # Zona horaria de Ecuador
        ecuador_tz = pytz.timezone('America/Guayaquil')
        utc_tz = pytz.utc

        # Si hora_ecuador no tiene zona horaria (naive), la localizamos
        if hour_ecuador.tzinfo is None:
            ecuador_time = ecuador_tz.localize(
                hour_ecuador)
        else:
            # Si ya tiene zona horaria, simplemente asumimos que es de Ecuador
            ecuador_time = hour_ecuador.astimezone(ecuador_tz)
        whitout_time_zone = ecuador_time.astimezone(utc_tz)
        return whitout_time_zone.replace(tzinfo=None)


    def convertir_a_hora_ecuador(self, hora_utc):
        # Zona horaria de Ecuador
        ecuador_tz = pytz.timezone('America/Guayaquil')
        utc_tz = pytz.utc
        utc_time = utc_tz.localize(
            hora_utc)
        whitout_time_zone = utc_time.astimezone(ecuador_tz)
        return whitout_time_zone.replace(tzinfo=None)



    def get_holidays_national(self, date_start, date_stop, employee_id):
        if isinstance(employee_id, int):
            employee_id = request.env['hr.employee'].sudo().browse(employee_id)

        national = request.env['resource.calendar.leaves'].sudo().search([
            ('type_of_leave_holiday', '=', 'national'),
            ('holiday_id', '=', False),
            ('date_from', '<=', date_start),
            ('date_to', '>=', date_stop),
        ])
        local = request.env['resource.calendar.leaves'].sudo().search([
            ('type_of_leave_holiday', '=', 'local'),
            ('holiday_id', '=', False),
            ('date_from', '<=', date_start),
            ('date_to', '>=', date_stop),
            ('city_id', '=', employee_id.department_id.city_id.id),
        ])

        return national, local


    def get_calendar (self, model_import,employee, current_date, type_of_resource ):
        (
            ranges_to_day,
            max_hours_for_schedule,
            is_extraordinary,
            is_especial_turn
        ) = (model_import.get_range_resource_calendar
            (
            employee,
            current_date,
            False,
            "history"
        ))
        if is_especial_turn:
            next_day = current_date + timedelta(days=1)
            (
                next_list,
                max_hours_for_schedule,
                is_extraordinary,
                is_especial_turn
            ) = (model_import.get_range_resource_calendar
                (
                employee,
                next_day,
                False,
                "history"
            ))
            ranges_to_day.extend(next_list)
            ranges_to_day = model_import.filtrar_turno_especial(ranges_to_day)

        if not ranges_to_day:
            if type_of_resource == 'employee':
                (
                    ranges_to_day,
                    max_hours_for_schedule,
                    is_extraordinary,
                    is_especial_turn
                ) = (model_import.get_range_resource_calendar
                    (
                    employee,
                    current_date,
                    False,
                    "employee"
                ))
                if is_especial_turn:
                    next_day = current_date + timedelta(days=1)
                    (
                        next_list,
                        max_hours_for_schedule,
                        is_extraordinary,
                        is_especial_turn
                    ) = (model_import.get_range_resource_calendar
                        (
                        employee,
                        next_day,
                        False,
                        "employee"
                    ))
                    ranges_to_day.extend(next_list)
                    ranges_to_day = model_import.filtrar_turno_especial(ranges_to_day)

            elif type_of_resource == 'departament':
                ranges_to_day, max_hours_for_schedule, is_especial_turn = model_import.get_range_resource_calendar_for_departament(
                    employee, current_date)
                if is_especial_turn:
                    next_day = current_date + timedelta(days=1)
                    next_list, max_hours_for_schedule, is_especial_turn = model_import.get_range_resource_calendar_for_departament(
                        employee,
                        next_day)
                    ranges_to_day.extend(next_list)
                    ranges_to_day = model_import.filtrar_turno_especial(ranges_to_day)

        return ranges_to_day

    #### Reporte general de pagos
    @http.route("/payroll/download_payslip_xlsx/<int:run_id>", type="http", auth="user")
    def download_payslip_xlsx(self, run_id, **kwargs):
        payslip_run = request.env["hr.payslip.run"].browse(run_id)


        if not payslip_run:
            return request.not_found()

        output = io.BytesIO()

        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        border_format = workbook.add_format(
            {
                "border": 1,
            }
        )
        worksheet = workbook.add_worksheet("Hojas de Pago")
        headers = [
            "Nº",
            "NRO DE CEDULA",

        ]
        header_indices = {header: idx for idx, header in enumerate(headers)}
        index_headers = 4
        for col_num, header in enumerate(headers):
            worksheet.write(index_headers, col_num, header, border_format)
        cell_values = {}

        def get_cell_value(row, col):
            return cell_values.get((row, col), 0)

        row = 5

        num_employee_before_month = 0
        num_employee_disassociate_month = 0
        num_employee_link_month = 0

        date_month_back = payslip_run.date_end - relativedelta(months=1)
        num_employeed = 1
        for payslip in payslip_run.slip_ids:
            worksheet.write(row, 0, num_employeed, border_format)
            worksheet.write(
                row, 1, payslip.employee_id.identification_id, border_format
            )

            payslip_lines = request.env["hr.payslip.line"].search_read(
                [("slip_id.employee_id", "=", payslip.employee_id.id),
                 ("slip_id", "=", payslip.id),
                 ("salary_rule_id.appears_on_payslip", "=", True)],
                ["name", "total", "quantity", "rate", "data_extra"],
                order="sequence asc"
            )
            for line in payslip_lines:
                line_name = line["name"]
                line_total = line["data_extra"] or line["total"]
                if line_name not in header_indices:
                    col_idx = len(headers)
                    headers.append(line_name)
                    header_indices[line_name] = col_idx
                    worksheet.write(index_headers, col_idx, line_name, border_format)
                else:
                    col_idx = header_indices[line_name]
                # existing_value = get_cell_value(row, col_idx)
                total_value =  line_total
                worksheet.write(row, col_idx, total_value, border_format)
                cell_values[(row, col_idx)] = total_value


            contrato = payslip.employee_id.contract_id

            # Validar empleados que laboraron en noviembre
            if contrato.date_start <= date_month_back and (
                    not contrato.date_end or contrato.date_end >= date_month_back
            ):
                num_employee_before_month += 1

            # Validar empleados que se desvinculan en diciembre
            if (
                    contrato.date_end
                    and contrato.date_end.month == payslip_run.date_end.month
                    and contrato.date_end.year == payslip_run.date_end.year
            ):
                num_employee_disassociate_month += 1

            # Validar empleados que se vinculan en diciembre
            if (
                    contrato.date_start
                    and contrato.date_start.month == payslip_run.date_end.month
                    and contrato.date_start.year == payslip_run.date_end.year
            ):
                num_employee_link_month += 1

            row += 1
            num_employeed += 1

        worksheet.conditional_format(
            5,
            0,
            row - 1,
            len(headers) - 1,
            {"type": "cell", "criteria": ">=", "value": 0, "format": border_format},
        )

        col_count = len(headers)
        header_format = workbook.add_format(
            {
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_size": 14,
                # 'bg_color': '#D7E4BC'
            }
        )

        # Escribir y combinar el encabezado
        worksheet.merge_range(
            0,
            0,
            0,
            col_count - 1,
            "FARMACIA CUXIBAMBA FARMACUX CIA LTDA",
            header_format,
        )
        worksheet.merge_range(
            1, 0, 1, col_count - 1, "RUC NRO. 1191751422001", header_format
        )
        worksheet.merge_range(
            2, 0, 2, col_count - 1, str(payslip.struct_id.name).upper().strip(), header_format
        )
        worksheet.merge_range(
            3,
            0,
            3,
            col_count - 1,
            f"{self.get_format_date_str(payslip_run.date_start, payslip_run.date_end)}",
            header_format,
        )

        footer_format = workbook.add_format(
            {
                "bold": True,
                "align": "center",
                "valign": "vcenter",
                "font_size": 12,
                # 'bg_color': '#D7E4BC'
                "border": 1,
            }
        )

        mes_en_ingles = date_month_back.strftime("%B")
        mes_en_espanol = meses_espanol.get(mes_en_ingles, mes_en_ingles)

        if payslip.struct_id.name == "Rol de Pagos":


            fecha_str = date_month_back.strftime(f" {mes_en_espanol} %Y")
            worksheet.merge_range(
                row + 2,
                0,
                row + 2,
                8,
                f"TOTAL DE TRABAJADORES MES DE {fecha_str}",
                footer_format,
            )
            worksheet.write(row + 2, 9, num_employee_before_month, border_format)

            mes_en_ingles2 = payslip_run.date_end.strftime("%B")
            mes_en_espanol2 = meses_espanol.get(mes_en_ingles2, mes_en_ingles2)

            fecha2_str = payslip_run.date_end.strftime(f" {mes_en_espanol2} %Y")
            worksheet.merge_range(
                row + 3,
                0,
                row + 3,
                8,
                f"PERSONAL QUE SE DESVINCULA EN {fecha2_str}",
                footer_format,
            )
            worksheet.write(row + 3, 9, num_employee_disassociate_month, border_format)

            worksheet.merge_range(
                row + 4,
                0,
                row + 4,
                8,
                f"PERSONAL QUE INGRESA A LABORAR EN EL MES DE {fecha2_str}",
                footer_format,
            )
            worksheet.write(row + 4, 9, num_employee_link_month, border_format)

            worksheet.merge_range(
                row + 5,
                0,
                row + 5,
                8,
                f"TOTAL DE TRABAJADORES ACTIVOS EN NONIMA {fecha2_str}",
                footer_format,
            )
            total = (
                    num_employee_before_month
                    + num_employee_disassociate_month
                    + num_employee_link_month
            )
            worksheet.write(row + 5, 9, total, border_format)

        workbook.close()
        output.seek(0)
        xlsx_data = output.read()
        output.close()

        return request.make_response(
            xlsx_data,
            headers=[
                ("Content-Type", "application/vnd.ms-excel"),
                ("Content-Disposition", content_disposition("hojas_de_pago.xlsx")),
            ],
        )

    def get_format_date_str(self, start_date, end_date):

        # Caso si las fechas son iguales (un solo día)
        if start_date == end_date:
            mes_en_ingles = start_date.strftime("%B")
            mes_en_espanol = meses_espanol.get(mes_en_ingles, mes_en_ingles)
            formatted_dates = start_date.strftime(f"%-d de {mes_en_espanol} de %Y")

        # Caso si las fechas están en el mismo mes y año
        elif start_date.month == end_date.month and start_date.year == end_date.year:
            mes_en_ingles = end_date.strftime("%B")
            mes_en_espanol = meses_espanol.get(mes_en_ingles, mes_en_ingles)
            formatted_dates = f'del {start_date.strftime("%-d")} al {end_date.strftime(f"%-d de {mes_en_espanol} de %Y")}'

        # Caso si las fechas están en el mismo año pero en meses diferentes
        elif start_date.year == end_date.year:
            mes_en_ingles = end_date.strftime("%B")
            mes_en_espanol = meses_espanol.get(mes_en_ingles, mes_en_ingles)
            formatted_dates = f'del {start_date.strftime("%-d de %B")} al {end_date.strftime(f"%-d de {mes_en_espanol} de %Y")}'

        # Caso si las fechas abarcan varios años
        else:
            mes_en_ingles = end_date.strftime("%B")
            mes_en_espanol = meses_espanol.get(mes_en_ingles, mes_en_ingles)
            mes_en_ingles_strart = start_date.strftime("%B")
            mes_en_espanol_start = meses_espanol.get(
                mes_en_ingles_strart, mes_en_ingles_strart
            )
            formatted_dates = f'del {start_date.strftime(f"%-d de {mes_en_espanol_start} de %Y")} al {end_date.strftime(f"%-d de {mes_en_espanol} de %Y")}'
        return formatted_dates



    ###### Reporte de empleados no econtrados o sin contrato

    @http.route('/reporte_empleados_sin_contrato/download_xlsx', type='http', auth='user')
    def download_employees_contract_report_xlsx(self, employees_data=None, missing_data=None, report_name='reporte',
                                                **kwargs):

        try:
            employees_without_contracts = []
            missing_identifications = []

            if employees_data:
                try:
                    employees_without_contracts = json.loads(base64.b64decode(employees_data).decode('utf-8'))
                except:
                    employees_without_contracts = []

            if missing_data:
                try:
                    missing_identifications = json.loads(base64.b64decode(missing_data).decode('utf-8'))
                except:
                    missing_identifications = []

            xlsx_buffer = io.BytesIO()
            try:
                workbook = xlsxwriter.Workbook(xlsx_buffer, {'in_memory': True})

                self._generate_employees_without_contracts_sheet(workbook, employees_without_contracts)
                self._generate_missing_employees_sheet(workbook, missing_identifications)

                workbook.close()
                xlsx_buffer.seek(0)

                filename = f'empleados_sin_contrato_{report_name}.xlsx'

                response = request.make_response(
                    xlsx_buffer.getvalue(),
                    headers=[
                        ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                        ('Content-Disposition', content_disposition(filename))
                    ]
                )

                return response

            finally:
                xlsx_buffer.close()

        except Exception as e:
            _logger.error(f"Error generando Excel de empleados sin contrato: {str(e)}")
            return request.make_response(f"Error interno del servidor: {str(e)}", status=500)

    def _generate_employees_without_contracts_sheet(self, workbook, employees_without_contracts):

        worksheet = workbook.add_worksheet('Empleados Sin Contrato')

        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#D7E4BC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })

        warning_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#FFC7CE',
            'font_color': '#9C0006',
            'bold': True
        })

        headers = ['ID Empleado', 'Identificación', 'Nombre', 'Estado']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        row = 1
        for employee in employees_without_contracts:
            worksheet.write(row, 0, employee.get('id', ''), cell_format)
            worksheet.write(row, 1, employee.get('identification_id', ''), cell_format)
            worksheet.write(row, 2, employee.get('name', ''), cell_format)
            worksheet.write(row, 3, 'REVISAR CONTRATO', warning_format)
            row += 1

        worksheet.set_column('A:A', 12)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 35)
        worksheet.set_column('D:D', 20)

        if employees_without_contracts:
            summary_format = workbook.add_format({
                'bold': True,
                'bg_color': '#FFE6CC',
                'border': 1
            })
            worksheet.merge_range(row + 1, 0, row + 1, 3,
                                  f'Total empleados sin contrato: {len(employees_without_contracts)}',
                                  summary_format)

    def _generate_missing_employees_sheet(self, workbook, missing_identifications):

        worksheet = workbook.add_worksheet('Empleados No Encontrados')

        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#F2DCDB',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })

        error_format = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'bg_color': '#F2DCDB',
            'font_color': '#A94442',
            'bold': True
        })

        headers = ['Identificación', 'Estado']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        row = 1
        for identification in missing_identifications:
            if identification:
                worksheet.write(row, 0, str(identification), cell_format)
                worksheet.write(row, 1, 'EMPLEADO NO ENCONTRADO', error_format)
                row += 1

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 25)

        if missing_identifications:
            summary_format = workbook.add_format({
                'bold': True,
                'bg_color': '#FFE6CC',
                'border': 1
            })
            worksheet.merge_range(row + 1, 0, row + 1, 1,
                                  f'Total empleados no encontrados: {len(missing_identifications)}',
                                  summary_format)

    def get_column_width(self, values, min_width=12):
        valor = max([len(str(v)) for v in values] + [min_width])
        return valor




