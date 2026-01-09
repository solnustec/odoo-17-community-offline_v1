# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, timedelta


class SurveyBranchVisit(models.Model):
    """
    Visita programada a sucursal para evaluación.
    Permite al administrativo programar visitas de empleados a sucursales
    para que realicen encuestas de evaluación.
    """
    _name = 'survey.branch.visit'
    _description = 'Visita Programada a Sucursal'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'scheduled_date asc, id desc'
    _rec_name = 'display_name'

    display_name = fields.Char(
        string='Nombre',
        compute='_compute_display_name',
        store=True
    )
    employee_id = fields.Many2one(
        'hr.employee',
        string='Empleado',
        required=True,
        tracking=True,
        index=True
    )
    branch_id = fields.Many2one(
        'hr.department',
        string='Sucursal',
        required=True,
        tracking=True,
        index=True,
        help="Departamento/Sucursal a visitar para la evaluación"
    )
    scheduled_date = fields.Date(
        string='Fecha Programada',
        required=True,
        tracking=True,
        index=True
    )
    survey_id = fields.Many2one(
        'survey.survey',
        string='Encuesta',
        required=True,
        tracking=True
    )
    state = fields.Selection([
        ('programada', 'Programada'),
        ('completada', 'Completada'),
        ('vencida', 'Vencida'),
        ('cancelada', 'Cancelada'),
    ], string='Estado', default='programada', tracking=True, index=True)

    user_input_id = fields.Many2one(
        'survey.user_input',
        string='Respuesta',
        readonly=True
    )
    scheduled_by_id = fields.Many2one(
        'res.users',
        string='Programado por',
        default=lambda self: self.env.user,
        readonly=True
    )
    notes = fields.Text(string='Notas')

    # Campos relacionados para facilitar búsquedas y reportes
    department_id = fields.Many2one(
        'hr.department',
        string='Departamento',
        related='employee_id.department_id',
        store=True,
        readonly=True
    )

    # Campos computados
    is_overdue = fields.Boolean(
        string='Vencida',
        compute='_compute_is_overdue',
        store=True
    )
    days_until = fields.Integer(
        string='Días para visita',
        compute='_compute_days_until'
    )

    @api.depends('employee_id', 'branch_id', 'scheduled_date')
    def _compute_display_name(self):
        for record in self:
            if record.employee_id and record.branch_id and record.scheduled_date:
                record.display_name = f"{record.employee_id.name} → {record.branch_id.name} ({record.scheduled_date})"
            elif record.employee_id and record.branch_id:
                record.display_name = f"{record.employee_id.name} → {record.branch_id.name}"
            else:
                record.display_name = "Nueva Visita"

    @api.depends('scheduled_date', 'state')
    def _compute_is_overdue(self):
        today = date.today()
        for record in self:
            record.is_overdue = (
                record.scheduled_date and
                record.scheduled_date < today and
                record.state == 'programada'
            )

    def _compute_days_until(self):
        today = date.today()
        for record in self:
            if record.scheduled_date:
                delta = (record.scheduled_date - today).days
                record.days_until = delta
            else:
                record.days_until = 0

    @api.constrains('scheduled_date')
    def _check_scheduled_date(self):
        for record in self:
            if record.scheduled_date and record.scheduled_date < date.today():
                # Permitir fechas pasadas pero con advertencia (para correcciones)
                pass

    @api.constrains('employee_id', 'branch_id', 'scheduled_date', 'survey_id')
    def _check_duplicate_visit(self):
        for record in self:
            existing = self.search([
                ('id', '!=', record.id),
                ('employee_id', '=', record.employee_id.id),
                ('branch_id', '=', record.branch_id.id),
                ('scheduled_date', '=', record.scheduled_date),
                ('survey_id', '=', record.survey_id.id),
                ('state', 'not in', ['cancelada']),
            ])
            if existing:
                raise ValidationError(_(
                    'Ya existe una visita programada de %s a %s el %s para esta encuesta.'
                ) % (record.employee_id.name, record.branch_id.name, record.scheduled_date))

    @api.model
    def create(self, vals):
        visit = super().create(vals)
        visit._send_notification()
        return visit

    def _send_notification(self):
        """Envía notificación al empleado sobre la visita programada."""
        self.ensure_one()
        employee = self.employee_id

        if employee and employee.user_id:
            # Crear actividad
            activity_type = self.env.ref('mail.mail_activity_data_todo')

            self.env['mail.activity'].create({
                'res_model_id': self.env['ir.model']._get_id('survey.branch.visit'),
                'res_id': self.id,
                'user_id': employee.user_id.id,
                'activity_type_id': activity_type.id,
                'date_deadline': self.scheduled_date,
                'summary': _('Visita programada: %s') % self.branch_id.name,
                'note': _('''<p>Tiene una visita programada para evaluar una sucursal.</p>
<p><strong>Sucursal:</strong> %s</p>
<p><strong>Fecha:</strong> %s</p>
<p><strong>Encuesta:</strong> %s</p>
''') % (self.branch_id.name, self.scheduled_date, self.survey_id.title),
            })

            # Enviar email si tiene correo
            if employee.work_email:
                template = self.env.ref('internal_control.email_template_branch_visit', raise_if_not_found=False)
                if template:
                    template.send_mail(self.id, force_send=True)

    def action_open_survey(self):
        """Abre la encuesta para responder, vinculando la sucursal."""
        self.ensure_one()
        survey = self.survey_id

        if not survey or not survey.access_token:
            raise UserError(_('La encuesta no tiene token de acceso público.'))

        if self.state == 'completada':
            raise UserError(_('Esta visita ya ha sido completada.'))

        if self.state == 'cancelada':
            raise UserError(_('Esta visita ha sido cancelada.'))

        # Buscar o crear user_input
        UserInput = self.env['survey.user_input']
        user_input = UserInput.search([
            ('survey_id', '=', survey.id),
            ('partner_id', '=', self.employee_id.user_id.partner_id.id),
            ('branch_visit_id', '=', self.id)
        ], limit=1)

        if not user_input:
            user_input = UserInput.create({
                'survey_id': survey.id,
                'partner_id': self.employee_id.user_id.partner_id.id if self.employee_id.user_id else False,
                'employee_id': self.employee_id.id,
                'branch_visit_id': self.id,
                'state': 'new',
                'email': self.employee_id.user_id.email if self.employee_id.user_id else self.employee_id.work_email,
                'is_admin_assigned': True,
                'assigned_by_id': self.env.user.id,
            })
            self.write({'user_input_id': user_input.id})

        url = f'/survey/start/{survey.access_token}?answer_token={user_input.access_token}&branch_visit_id={self.id}'

        return {
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }

    def action_view_response(self):
        """Ver la respuesta de la encuesta."""
        self.ensure_one()
        if not self.user_input_id:
            raise UserError(_('No hay respuesta asociada a esta visita.'))

        return {
            'type': 'ir.actions.act_window',
            'name': _('Respuesta'),
            'res_model': 'survey.user_input',
            'res_id': self.user_input_id.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def action_cancel(self):
        """Cancelar la visita."""
        for record in self:
            if record.state == 'completada':
                raise UserError(_('No se puede cancelar una visita ya completada.'))
            record.write({'state': 'cancelada'})

    @api.model
    def _cron_check_overdue_visits(self):
        """Cron job para marcar visitas vencidas."""
        today = date.today()
        overdue_visits = self.search([
            ('scheduled_date', '<', today),
            ('state', '=', 'programada')
        ])
        overdue_visits.write({'state': 'vencida'})


class SurveyBranchVisitWizard(models.TransientModel):
    """
    Wizard para programar visitas en lote.
    Permite al administrativo agregar múltiples visitas de una sola vez.
    """
    _name = 'survey.branch.visit.wizard'
    _description = 'Programar Visitas en Lote'

    survey_id = fields.Many2one(
        'survey.survey',
        string='Encuesta',
        required=True
    )
    line_ids = fields.One2many(
        'survey.branch.visit.wizard.line',
        'wizard_id',
        string='Visitas a Programar'
    )
    visit_count = fields.Integer(
        string='Total Visitas',
        compute='_compute_visit_count'
    )

    # Campos para importación
    import_file = fields.Binary(string='Archivo de Importación')
    import_filename = fields.Char(string='Nombre del Archivo')
    import_result = fields.Text(string='Resultado de Importación', readonly=True)

    @api.depends('line_ids')
    def _compute_visit_count(self):
        for wizard in self:
            wizard.visit_count = len(wizard.line_ids.filtered(
                lambda l: l.employee_id and l.branch_id and l.scheduled_date
            ))

    def action_download_template(self):
        """Descargar plantilla XLSX para importación."""
        self.ensure_one()
        import base64
        import io

        try:
            import xlsxwriter
        except ImportError:
            raise UserError(_('La librería xlsxwriter no está instalada.'))

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Visitas')

        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
        })
        instruction_format = workbook.add_format({
            'italic': True,
            'font_color': '#666666',
            'text_wrap': True,
        })
        example_format = workbook.add_format({
            'bg_color': '#E2EFDA',
            'border': 1,
        })

        # Encabezados
        headers = ['empleado_identificacion', 'sucursal_codigo', 'sucursal_nombre', 'fecha_programada', 'notas']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Ejemplo 1 - con código de departamento
        worksheet.write(1, 0, '1234567890', example_format)
        worksheet.write(1, 1, '1.46', example_format)
        worksheet.write(1, 2, '', example_format)
        worksheet.write(1, 3, '2026-01-15', example_format)
        worksheet.write(1, 4, 'Visita de evaluación', example_format)

        # Ejemplo 2 - con nombre
        worksheet.write(2, 0, '0987654321', example_format)
        worksheet.write(2, 1, '', example_format)
        worksheet.write(2, 2, 'Amaluza', example_format)
        worksheet.write(2, 3, '2026-01-16', example_format)
        worksheet.write(2, 4, '', example_format)

        # Instrucciones en hoja separada
        instructions = workbook.add_worksheet('Instrucciones')
        title_format = workbook.add_format({'bold': True, 'font_size': 14})
        bold_format = workbook.add_format({'bold': True})

        instructions.write(0, 0, 'INSTRUCCIONES PARA IMPORTAR VISITAS', title_format)
        instructions.write(2, 0, '1. Complete la hoja "Visitas" con los datos de las visitas a programar.')
        instructions.write(3, 0, '2. Elimine las filas de ejemplo (filas 2 y 3) antes de importar.')
        instructions.write(4, 0, '')
        instructions.write(5, 0, 'COLUMNAS:', bold_format)
        instructions.write(6, 0, '• empleado_identificacion: Número de cédula/identificación del empleado')
        instructions.write(7, 0, '')
        instructions.write(8, 0, 'SUCURSAL (use código O nombre, no es necesario ambos):', bold_format)
        instructions.write(9, 0, '• sucursal_codigo: Código de departamento/sucursal (ej: 1.46). Tiene prioridad si se llenan ambos.')
        instructions.write(10, 0, '• sucursal_nombre: Nombre de la sucursal (ej: Amaluza)')
        instructions.write(11, 0, '  → Puede usar solo el código, solo el nombre, o ambos (el código tiene prioridad)')
        instructions.write(12, 0, '')
        instructions.write(13, 0, '• fecha_programada: Fecha en formato YYYY-MM-DD (ej: 2026-01-15)')
        instructions.write(14, 0, '• notas: Notas opcionales para la visita')
        instructions.set_column(0, 0, 80)

        # Ajustar anchos de columna
        worksheet.set_column(0, 0, 25)
        worksheet.set_column(1, 1, 18)
        worksheet.set_column(2, 2, 25)
        worksheet.set_column(3, 3, 18)
        worksheet.set_column(4, 4, 30)

        workbook.close()
        output.seek(0)

        # Crear adjunto
        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_visitas.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_import_file(self):
        """Importar visitas desde archivo CSV/Excel."""
        self.ensure_one()
        import base64
        import csv
        import io

        if not self.import_file:
            raise UserError(_('Seleccione un archivo para importar.'))

        file_content = base64.b64decode(self.import_file)
        filename = self.import_filename or ''

        lines_data = []
        errors = []

        if filename.lower().endswith('.csv'):
            try:
                csv_content = file_content.decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                for row_num, row in enumerate(csv_reader, 2):
                    # Saltar líneas de comentario
                    if any(str(v).startswith('#') for v in row.values()):
                        continue
                    lines_data.append((row_num, row))
            except Exception as e:
                raise UserError(_('Error al leer archivo CSV: %s') % str(e))

        elif filename.lower().endswith('.xlsx'):
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = workbook.active
                headers = [str(cell.value or '').strip().lower() for cell in sheet[1]]

                for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), 2):
                    row = {}
                    for col_idx, cell in enumerate(row_cells):
                        if col_idx < len(headers):
                            row[headers[col_idx]] = cell.value or ''
                    # Saltar filas vacías
                    if not any(row.values()):
                        continue
                    lines_data.append((row_idx, row))
            except ImportError:
                raise UserError(_('Para importar archivos Excel, instale la librería openpyxl.'))
            except Exception as e:
                raise UserError(_('Error al leer archivo Excel: %s') % str(e))

        elif filename.lower().endswith('.xls'):
            try:
                import xlrd
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)
                headers = [str(sheet.cell_value(0, col)).strip().lower() for col in range(sheet.ncols)]

                for row_idx in range(1, sheet.nrows):
                    row = {}
                    for col_idx, header in enumerate(headers):
                        row[header] = sheet.cell_value(row_idx, col_idx)
                    # Saltar filas vacías
                    if not any(row.values()):
                        continue
                    lines_data.append((row_idx + 1, row))
            except ImportError:
                raise UserError(_('Para importar archivos .xls, instale la librería xlrd.'))
            except Exception as e:
                raise UserError(_('Error al leer archivo Excel: %s') % str(e))
        else:
            raise UserError(_('Formato no soportado. Use CSV o Excel (.xlsx, .xls)'))

        # Procesar las líneas
        Employee = self.env['hr.employee']
        Department = self.env['hr.department']
        WizardLine = self.env['survey.branch.visit.wizard.line']

        created_lines = 0
        for row_num, row in lines_data:
            try:
                # Buscar empleado por identificación o nombre
                emp_id = str(row.get('empleado_identificacion', '')).strip()
                employee = Employee.search([
                    '|',
                    ('identification_id', '=', emp_id),
                    ('name', 'ilike', emp_id)
                ], limit=1)

                if not employee:
                    errors.append(f"Fila {row_num}: Empleado '{emp_id}' no encontrado")
                    continue

                # Buscar sucursal (departamento) por código o nombre (prioridad al código)
                branch_code = str(row.get('sucursal_codigo', '')).strip()
                branch_name = str(row.get('sucursal_nombre', '')).strip()
                branch = None

                # Validar que al menos uno esté presente
                if not branch_code and not branch_name:
                    errors.append(f"Fila {row_num}: Debe proporcionar sucursal_codigo o sucursal_nombre (ambos están vacíos)")
                    continue

                # Primero intentar por código si está presente
                if branch_code:
                    branch = Department.search([('code', '=', branch_code)], limit=1)
                    if not branch:
                        # Intentar búsqueda parcial por código
                        branch = Department.search([('code', 'ilike', branch_code)], limit=1)

                # Si no se encontró por código, intentar por nombre
                if not branch and branch_name:
                    branch = Department.search([('name', '=', branch_name)], limit=1)
                    if not branch:
                        # Intentar búsqueda parcial por nombre
                        branch = Department.search([('name', 'ilike', branch_name)], limit=1)

                # Si no hay nombre pero hay código, intentar buscar el código como nombre
                if not branch and branch_code and not branch_name:
                    branch = Department.search([('name', 'ilike', branch_code)], limit=1)

                if not branch:
                    if branch_code and branch_name:
                        errors.append(f"Fila {row_num}: Sucursal no encontrada. Código '{branch_code}' y nombre '{branch_name}' no coinciden con ninguna sucursal")
                    elif branch_code:
                        errors.append(f"Fila {row_num}: Sucursal con código '{branch_code}' no encontrada")
                    else:
                        errors.append(f"Fila {row_num}: Sucursal con nombre '{branch_name}' no encontrada")
                    continue

                # Parsear fecha
                date_str = str(row.get('fecha_programada', '')).strip()
                try:
                    scheduled_date = fields.Date.from_string(date_str)
                except:
                    errors.append(f"Fila {row_num}: Fecha inválida '{date_str}'. Use formato YYYY-MM-DD")
                    continue

                # Crear línea
                WizardLine.create({
                    'wizard_id': self.id,
                    'employee_id': employee.id,
                    'branch_id': branch.id,
                    'scheduled_date': scheduled_date,
                    'notes': str(row.get('notas', '')).strip() or False,
                })
                created_lines += 1

            except Exception as e:
                errors.append(f"Fila {row_num}: Error - {str(e)}")

        # Construir mensaje de resultado
        total_processed = len(lines_data)
        failed_count = len(errors)

        result_lines = []
        result_lines.append(_('=== RESUMEN DE IMPORTACIÓN ==='))
        result_lines.append(_('Filas procesadas: %d') % total_processed)
        result_lines.append(_('Importadas exitosamente: %d') % created_lines)
        result_lines.append(_('Errores: %d') % failed_count)

        if errors:
            result_lines.append('')
            result_lines.append(_('=== ERRORES ENCONTRADOS ==='))
            # Mostrar todos los errores (hasta 20)
            for error in errors[:20]:
                result_lines.append('• ' + error)
            if len(errors) > 20:
                result_lines.append(_('... y %d errores más') % (len(errors) - 20))

        if created_lines == 0 and total_processed > 0:
            result_lines.append('')
            result_lines.append(_('⚠ No se importó ninguna línea. Revise los errores arriba.'))
        elif created_lines > 0:
            result_lines.append('')
            result_lines.append(_('✓ Las líneas importadas aparecen en la tabla de abajo.'))

        import_result = '\n'.join(result_lines)

        # Limpiar archivo y guardar resultado
        self.write({
            'import_file': False,
            'import_filename': False,
            'import_result': import_result
        })

        # Recargar wizard para mostrar resultado
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'survey.branch.visit.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_schedule_visits(self):
        """Crear las visitas programadas."""
        self.ensure_one()

        valid_lines = self.line_ids.filtered(
            lambda l: l.employee_id and l.branch_id and l.scheduled_date
        )

        if not valid_lines:
            raise UserError(_('Debe agregar al menos una visita con todos los campos completos.'))

        Visit = self.env['survey.branch.visit']
        created_visits = Visit
        errors = []

        for idx, line in enumerate(valid_lines, 1):
            try:
                visit = Visit.create({
                    'employee_id': line.employee_id.id,
                    'branch_id': line.branch_id.id,
                    'scheduled_date': line.scheduled_date,
                    'survey_id': self.survey_id.id,
                    'notes': line.notes,
                })
                created_visits |= visit
            except ValidationError as e:
                errors.append(f"Línea {idx}: {str(e)}")

        if errors:
            raise UserError(_('Errores al crear visitas:\n') + '\n'.join(errors))

        # Retornar acción para ver las visitas creadas
        return {
            'type': 'ir.actions.act_window',
            'name': _('Visitas Programadas (%d)') % len(created_visits),
            'res_model': 'survey.branch.visit',
            'view_mode': 'tree,form',
            'domain': [('id', 'in', created_visits.ids)],
            'target': 'current',
        }


class SurveyBranchVisitWizardLine(models.TransientModel):
    """Línea del wizard para programar visitas."""
    _name = 'survey.branch.visit.wizard.line'
    _description = 'Línea de Wizard de Visitas'

    wizard_id = fields.Many2one(
        'survey.branch.visit.wizard',
        string='Wizard',
        required=True,
        ondelete='cascade'
    )
    employee_id = fields.Many2one(
        'hr.employee',
        string='Empleado'
    )
    branch_id = fields.Many2one(
        'hr.department',
        string='Sucursal'
    )
    scheduled_date = fields.Date(
        string='Fecha'
    )
    notes = fields.Char(
        string='Notas'
    )

    def action_duplicate(self):
        """Duplicar esta línea."""
        self.ensure_one()
        self.copy({
            'wizard_id': self.wizard_id.id,
            'employee_id': self.employee_id.id,
            'branch_id': False,
            'scheduled_date': False,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'survey.branch.visit.wizard',
            'res_id': self.wizard_id.id,
            'view_mode': 'form',
            'target': 'new',
        }
