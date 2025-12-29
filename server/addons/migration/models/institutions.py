from datetime import date
from odoo import models, fields, api
from odoo.exceptions import ValidationError
from odoo.exceptions import UserError
import base64
import io
import openpyxl


class Institution(models.Model):
    _name = 'institution'
    _description = 'Institution'

    id_institutions = fields.Char(string='ID.Institucion', required=True)
    name = fields.Char(string='Nombre', required=True)
    ruc_institution = fields.Char(string='RUC', required=False)
    agreement_date = fields.Date(string='Fecha del Convenio')
    address = fields.Char(string='Dirección')
    type_credit_institution = fields.Selection(
        selection=[
            ('discount', 'Descuento'),
            ('credit', 'Crédito')
        ],
        string='Tipo de Institución',
        required=True,
        help='Seleccione si la institución opera con descuentos o con crédito.'
    )
    cellphone = fields.Char(string='Teléfono')
    court_day = fields.Integer(string='Día de Corte (Créditos)')
    additional_discount_percentage = fields.Float(
        string='Descuento Adicional (%)',
        help='Porcentaje adicional aplicado al total de la factura.'
    )
    pvp = fields.Selection(
        [('1', 'Activada'), ('0', 'Desactivada')],
        string='Estado (PVP)',
        default='1',
        help='Estado de la institución. 1: Activada, 0: Desactivada'
    )
    pos_ids = fields.Many2many(
        'pos.config',
        string='Puntos de Venta',
        help='Seleccione los puntos de venta asociados a esta institución.'
    )
    institution_client_ids = fields.One2many(
        'institution.client',
        'institution_id',
        string='Clientes Asociados',
        help='Lista de clientes asociados a esta institución con sus montos.'
    )
    excel_file = fields.Binary(string="Archivo Excel")
    excel_filename = fields.Char(string="Nombre del Archivo")

    def action_import_clients_excel(self):
        """Procesa el Excel y crea/actualiza/elimina clientes de la institución usando RUC/Cédula."""
        import base64
        import io
        import openpyxl
        from odoo.exceptions import UserError

        for record in self:
            if not record.excel_file:
                raise UserError("Por favor, suba un archivo Excel antes de importar.")

            file_data = base64.b64decode(record.excel_file)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]
            expected = ['institution_id', 'partner_vat', 'sale', 'available_amount', 'action']

            # Validar cabecera
            if headers != expected:
                raise UserError(f"Las columnas deben ser: {', '.join(expected)}")

            inst_client_model = self.env['institution.client']

            for row in sheet.iter_rows(min_row=2, values_only=True):
                institution_id, partner_vat, sale, available_amount, action = row

                if not institution_id or not partner_vat or not action:
                    continue

                # Buscar institución
                institution = self.env['institution'].browse(int(institution_id))
                if not institution:
                    raise UserError(f"No se encontró institución con ID {institution_id}")

                # Buscar cliente por RUC/Cédula
                partner = self.env['res.partner'].search([('vat', '=', str(partner_vat))], limit=1)
                if not partner:
                    raise UserError(f"No se encontró cliente con RUC/Cédula {partner_vat}")

                if action == 'add':
                    existing = inst_client_model.search([
                        ('institution_id', '=', institution.id),
                        ('partner_id', '=', partner.id)
                    ])
                    if existing:
                        existing.write({
                            'sale': sale,
                            'available_amount': available_amount
                        })
                    else:
                        inst_client_model.create({
                            'institution_id': institution.id,
                            'partner_id': partner.id,
                            'sale': sale,
                            'available_amount': available_amount
                        })

                elif action == 'update':
                    client_link = inst_client_model.search([
                        ('institution_id', '=', institution.id),
                        ('partner_id', '=', partner.id)
                    ])
                    if not client_link:
                        raise UserError(f"No existe vínculo entre {partner.name} y {institution.name}")
                    client_link.write({
                        'sale': sale,
                        'available_amount': available_amount
                    })

                elif action == 'delete':
                    client_link = inst_client_model.search([
                        ('institution_id', '=', institution.id),
                        ('partner_id', '=', partner.id)
                    ])
                    client_link.unlink()

                else:
                    raise UserError(f"Acción desconocida '{action}' en la fila con RUC/Cédula {partner_vat}")

            workbook.close()

    def action_export_clients_excel(self):
        """Exporta los clientes asociados a la institución actual a un archivo Excel."""
        import io
        import base64
        import openpyxl
        from openpyxl.utils import get_column_letter
        from odoo.exceptions import UserError

        self.ensure_one()

        # Validar que haya clientes asociados
        clients = self.institution_client_ids
        if not clients:
            raise UserError("No hay clientes asociados para exportar.")

        # Crear workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Clientes Institución"

        headers = ['institution_id', 'partner_vat', 'partner_name', 'sale', 'available_amount']
        sheet.append(headers)

        for client in clients:
            sheet.append([
                client.institution_id.id,
                client.partner_id.vat or '',
                client.partner_id.name or '',
                client.sale or 0.0,
                client.available_amount or 0.0,
            ])

        # Ajustar ancho de columnas
        for i, column in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(i)].width = 22

        # Guardar en memoria
        fp = io.BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = base64.b64encode(fp.read())
        fp.close()

        # Crear archivo binario temporal para descargar
        export_name = f"Clientes_{self.name.replace(' ', '_')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': export_name,
            'type': 'binary',
            'datas': data,
            'res_model': 'institution',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        download_url = f'/web/content/{attachment.id}?download=true'
        return {
            'type': 'ir.actions.act_url',
            'url': download_url,
            'target': 'new',
        }

    def renew_credits_on_cutoff_day(self):
        today = date.today()
        for institution in self.search([]):
            if institution.court_day and today.day == institution.court_day:
                for client in institution.institution_client_ids:
                    new_credit_value = client.sale
                    client.available_amount = new_credit_value


class InstitutionClient(models.Model):
    _name = 'institution.client'
    _description = 'Relación entre Instituciones y Clientes con Montos'

    institution_id = fields.Many2one(
        'institution',
        string='Institución',
        required=True,
        help='Institución asociada al cliente.'
    )
    partner_id = fields.Many2one(
        'res.partner',
        string='Cliente',
        required=True,
        help='Cliente asociado a la institución.'
    )
    available_amount = fields.Float(
        string='Saldo',
        required=True,
        help='Monto asignado al cliente para esta institución.'
    )
    sale = fields.Float(
        string='Cupo asignado',
        required=True,
        help='Saldo total de clientes para esta institución.'
    )

    @api.onchange('sale')
    def _onchange_sale(self):
        if not self.id:
            self.available_amount = self.sale

    @api.onchange('available_amount')
    def _onchange_available_amount(self):
        print("valor do saldo: ", self.available_amount)
        print("valor da compra: ", self.sale)
        if self.available_amount < 0:
            raise ValidationError('El monto disponible no puede ser negativo.')
        if self.available_amount > self.sale:
            raise ValidationError('El monto disponible no puede ser mayor que el saldo total.')

    @api.model
    def create(self, vals):
        if 'sale' in vals and 'available_amount' not in vals:
            vals['available_amount'] = vals['sale']
        return super(InstitutionClient, self).create(vals)

    @api.model
    def get_institutions_by_partner(self, partner_id):
        institution_clients = self.search([('partner_id', '=', partner_id)])
        result = []

        for record in institution_clients:
            inst = record.institution_id
            if inst.type_credit_institution == 'credit':
                result.append({
                    'institution_id': inst.id,
                    'institution_name': inst.name,
                    'available_amount': record.available_amount,
                    'sale': record.sale,
                    'type_credit_institution': inst.type_credit_institution,
                    'cut_off_date': inst.court_day,
                })
        return result

    '''
    @api.constrains('available_amount', 'sale')
    def _check_available_amount(self):
        print("self: ", self)
        for record in self:
            print("valor do saldo 1: ", self.available_amount)
            print("valor da compra 1: ", self.sale)
            if record.available_amount < 0:
                raise ValidationError('El monto disponible no puede ser negativo.')
            if record.available_amount > record.sale:
                raise ValidationError('El monto disponible no puede ser mayor que el saldo total.')
    '''
